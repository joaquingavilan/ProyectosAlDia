from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View, ListView
from .models import *
from .forms import *
from django.db.models import Q, F, Sum, Exists, OuterRef
from django.conf import settings
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm, UserChangeForm
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import re, locale, json, os
from datetime import date, datetime
from io import BytesIO
from django.contrib import messages
from django.forms import formset_factory, BooleanField
from django.http import JsonResponse, HttpResponse, FileResponse, HttpResponseRedirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from django.core.files import File
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal
from django.core import serializers
from django.core.serializers import serialize
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.forms import PasswordChangeForm
from reportlab.pdfgen import canvas
import logging
from collections import defaultdict
from django.urls import reverse
from datetime import timedelta
from django.contrib.auth.forms import PasswordChangeForm
from django.core.files.storage import FileSystemStorage
from django.utils.text import get_valid_filename
from django.db.models import F, Case, When, IntegerField, Value
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import plotly
import plotly.graph_objs as go
import plotly.io as pio
from django.utils.safestring import mark_safe
import io
import urllib, base64
import base64
from io import BytesIO
import pandas as pd
import plotly.express as px
from datetime import datetime
from django.core.serializers.json import DjangoJSONEncoder


logger = logging.getLogger(__name__)

from django.db import connection


def plot_estado_ingenieros_json(ingenieros):
    data = {
        'nombres': [ingeniero.nombre_ingeniero for ingeniero in ingenieros],
        'obras_ejecucion': [ingeniero.cantidad_obras_ejecucion for ingeniero in ingenieros],
        'presupuestos_elaboracion': [ingeniero.cantidad_presupuestos_elaboracion for ingeniero in ingenieros],
    }
    return data


def plot_certificados_pendientes():
    certificados = HechoCertificado.objects.filter(fecha_envio__isnull=False, estado='S')
    proyectos_dict = {proyecto.proyecto_id: proyecto.nombre for proyecto in DimProyecto.objects.all()}
    presupuestos_dict = {presupuesto.id: presupuesto.proyecto_id for presupuesto in HechoPresupuesto.objects.all()}

    # Reestructurar datos para manejar múltiples certificados por proyecto
    proyectos_certificados = {}
    for certificado in certificados:
        presupuesto_id = certificado.presupuesto_id
        proyecto_id = presupuestos_dict[presupuesto_id]
        proyecto_nombre = proyectos_dict[proyecto_id]

        if proyecto_nombre not in proyectos_certificados:
            proyectos_certificados[proyecto_nombre] = []

        proyectos_certificados[proyecto_nombre].append({
            'id': certificado.id,
            'monto': float(certificado.monto_total)
        })

    graph_data = {
        'proyectos': list(proyectos_certificados.keys()),
        'certificados_proyecto': proyectos_certificados
    }

    return graph_data




def actualizar_dimensiones():
    with connection.cursor() as cursor:
        # Realizamos todas las consultas primero
        cursor.execute("""
            SELECT 
                m.id, m.nombre, m.marca_id, m.medida_id, 
                m.minimo, m.unidades_stock, m.fotografia, m.id_proveedor_id
            FROM 
                public."Aplicacion_material" m
        """)
        materiales = cursor.fetchall()

        cursor.execute("""
            SELECT 
                c.id, c.nombre, c.ruc, c.email, c.ciudad_id, c.direccion
            FROM 
                public."Aplicacion_cliente" c
        """)
        clientes = cursor.fetchall()

        cursor.execute("""
            SELECT 
                c.id, c.nombre
            FROM 
                public."Aplicacion_ciudad" c
        """)
        ciudades = cursor.fetchall()

        cursor.execute("""
            SELECT 
                u.id, u.username, u.first_name, u.last_name, u.email, 
                u.is_staff, u.is_active, u.date_joined
            FROM 
                public.auth_user u
        """)
        usuarios = cursor.fetchall()

        cursor.execute("""
            SELECT 
                p.id, p.nombre, p.ruc, p.email, p.ciudad_id, 
                p.direccion, p.pagina_web, p.observaciones, p.telefono
            FROM 
                public."Aplicacion_proveedor" p
        """)
        proveedores = cursor.fetchall()

        cursor.execute("""
            SELECT 
                p.id, p.nombre, p.cliente_id, p.ciudad_id
            FROM 
                public."Aplicacion_proyecto" p
        """)
        proyectos = cursor.fetchall()

        # Iniciamos una transacción
        with connection.cursor() as cursor:
            cursor.execute('BEGIN;')

            # Truncar tablas de dimensiones
            cursor.execute('TRUNCATE TABLE dim_material RESTART IDENTITY CASCADE;')
            cursor.execute('TRUNCATE TABLE dim_cliente RESTART IDENTITY CASCADE;')
            cursor.execute('TRUNCATE TABLE dim_ciudad RESTART IDENTITY CASCADE;')
            cursor.execute('TRUNCATE TABLE dim_usuario RESTART IDENTITY CASCADE;')
            cursor.execute('TRUNCATE TABLE dim_proveedor RESTART IDENTITY CASCADE;')
            cursor.execute('TRUNCATE TABLE dim_proyecto RESTART IDENTITY CASCADE;')

            # Insertar o actualizar materiales
            for material in materiales:
                material_id, nombre, marca_id, medida_id, minimo, unidades_stock, fotografia, id_proveedor_id = material
                cursor.execute("""
                    INSERT INTO dim_material (
                        material_id, nombre, marca_id, medida_id, minimo, 
                        unidades_stock, fotografia, id_proveedor_id
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (material_id) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        marca_id = EXCLUDED.marca_id,
                        medida_id = EXCLUDED.medida_id,
                        minimo = EXCLUDED.minimo,
                        unidades_stock = EXCLUDED.unidades_stock,
                        fotografia = EXCLUDED.fotografia,
                        id_proveedor_id = EXCLUDED.id_proveedor_id;
                """, [
                    material_id, nombre, marca_id, medida_id, minimo,
                    unidades_stock, fotografia, id_proveedor_id
                ])

            # Insertar o actualizar clientes
            for cliente in clientes:
                cliente_id, nombre, ruc, email, ciudad_id, direccion = cliente
                cursor.execute("""
                    INSERT INTO dim_cliente (
                        cliente_id, nombre, ruc, email, ciudad_id, direccion
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (cliente_id) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        ruc = EXCLUDED.ruc,
                        email = EXCLUDED.email,
                        ciudad_id = EXCLUDED.ciudad_id,
                        direccion = EXCLUDED.direccion;
                """, [
                    cliente_id, nombre, ruc, email, ciudad_id, direccion
                ])

            # Insertar o actualizar ciudades
            for ciudad in ciudades:
                ciudad_id, nombre = ciudad
                cursor.execute("""
                    INSERT INTO dim_ciudad (
                        ciudad_id, nombre
                    ) 
                    VALUES (%s, %s)
                    ON CONFLICT (ciudad_id) DO UPDATE SET
                        nombre = EXCLUDED.nombre;
                """, [
                    ciudad_id, nombre
                ])

            # Insertar o actualizar usuarios
            for usuario in usuarios:
                usuario_id, username, first_name, last_name, email, is_staff, is_active, date_joined = usuario
                cursor.execute("""
                    INSERT INTO dim_usuario (
                        usuario_id, username, first_name, last_name, email, 
                        is_staff, is_active, date_joined
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (usuario_id) DO UPDATE SET
                        username = EXCLUDED.username,
                        first_name = EXCLUDED.first_name,
                        last_name = EXCLUDED.last_name,
                        email = EXCLUDED.email,
                        is_staff = EXCLUDED.is_staff,
                        is_active = EXCLUDED.is_active,
                        date_joined = EXCLUDED.date_joined;
                """, [
                    usuario_id, username, first_name, last_name, email,
                    is_staff, is_active, date_joined
                ])

            # Insertar o actualizar proveedores
            for proveedor in proveedores:
                proveedor_id, nombre, ruc, email, ciudad_id, direccion, pagina_web, observaciones, telefono = proveedor
                cursor.execute("""
                    INSERT INTO dim_proveedor (
                        proveedor_id, nombre, ruc, email, ciudad_id, direccion, 
                        pagina_web, observaciones, telefono
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (proveedor_id) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        ruc = EXCLUDED.ruc,
                        email = EXCLUDED.email,
                        ciudad_id = EXCLUDED.ciudad_id,
                        direccion = EXCLUDED.direccion,
                        pagina_web = EXCLUDED.pagina_web,
                        observaciones = EXCLUDED.observaciones,
                        telefono = EXCLUDED.telefono;
                """, [
                    proveedor_id, nombre, ruc, email, ciudad_id, direccion,
                    pagina_web, observaciones, telefono
                ])

            # Insertar o actualizar proyectos
            for proyecto in proyectos:
                proyecto_id, nombre, cliente_id, ciudad_id = proyecto
                cursor.execute("""
                    INSERT INTO dim_proyecto (
                        proyecto_id, nombre, cliente_id, ciudad_id
                    ) 
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (proyecto_id) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        cliente_id = EXCLUDED.cliente_id,
                        ciudad_id = EXCLUDED.ciudad_id;
                """, [
                    proyecto_id, nombre, cliente_id, ciudad_id
                ])

            cursor.execute('COMMIT;')


def actualizar_hechos():
    with connection.cursor() as cursor:
        # Bloque de consultas

        # Consulta para HechoEstadoIngeniero
        cursor.execute("""
            SELECT 
                u.id AS ingeniero_id,
                u.username AS nombre_ingeniero,
                COALESCE(obras_en_ejecucion.cantidad, 0) AS cantidad_obras_ejecucion,
                COALESCE(presupuestos_en_elaboracion.cantidad, 0) AS cantidad_presupuestos_elaboracion
            FROM public.auth_user u
            LEFT JOIN (
                SELECT 
                    encargado_id,
                    COUNT(*) AS cantidad
                FROM public."Aplicacion_obra"
                WHERE estado = 'E'
                GROUP BY encargado_id
            ) obras_en_ejecucion ON u.id = obras_en_ejecucion.encargado_id
            LEFT JOIN (
                SELECT 
                    encargado_id,
                    COUNT(*) AS cantidad
                FROM public."Aplicacion_presupuesto"
                WHERE estado = 'E'
                GROUP BY encargado_id
            ) presupuestos_en_elaboracion ON u.id = presupuestos_en_elaboracion.encargado_id
            WHERE u.id IN (
                SELECT user_id
                FROM public.auth_user_groups
                WHERE group_id = (SELECT id FROM public.auth_group WHERE name = 'INGENIERO')
            );
        """)
        hechos_estado_ingeniero = cursor.fetchall()

        # Consulta para HechoCertificado
        cursor.execute("""
            SELECT 
                c.id, c.ingeniero_id, c.presupuesto_id, c.estado,
                c.fecha_creacion, c.fecha_envio, c.fecha_pago, c.iva,
                c.monto_total, c.subtotal
            FROM public."Aplicacion_certificado" c
        """)
        hechos_certificado = cursor.fetchall()

        # Consulta para HechoPresupuesto
        cursor.execute("""
            SELECT 
                p.id, p.monto_total, p.estado, p.encargado_id, p.proyecto_id, p.anticipo,
                p.monto_anticipo, p.fecha_pago_anticipo, p.comprobante_anticipo,
                p.iva, p.subtotal
            FROM public."Aplicacion_presupuesto" p
        """)
        hechos_presupuesto = cursor.fetchall()
        # Consulta para obtener la cantidad total de materiales por obra
        cursor.execute("""
            SELECT
                p.obra_id,
                mp.material_id,
                SUM(mp.cantidad) AS cantidad_total,
                pr.nombre AS obra_nombre,
                m.nombre AS material_nombre
            FROM
                public."Aplicacion_pedido" p
            JOIN
                public."Aplicacion_materialpedido" mp ON p.id = mp.pedido_id
            JOIN
                public."Aplicacion_obra" o ON p.obra_id = o.id
            JOIN
                public."Aplicacion_proyecto" pr ON o.proyecto_id = pr.id
            JOIN
                public."Aplicacion_material" m ON mp.material_id = m.id
            GROUP BY
                p.obra_id, mp.material_id, pr.nombre, m.nombre;
        """)
        hechos_materiales = cursor.fetchall()
        # Consulta para HechoCronograma
        cursor.execute("""
            SELECT 
                c.id AS cronograma_id,
                dc.id AS detalle_id,
                dc.fecha_programada,
                dc.fecha_culminacion
            FROM public."Aplicacion_cronograma" c
            JOIN public."Aplicacion_detallecronograma" dc ON c.id = dc.cronograma_id
        """)
        hechos_cronograma = cursor.fetchall()

        cursor.execute("""
            SELECT 
                o.id AS id_obra,
                o.encargado_id AS id_encargado,
                o.estado,
                o.fecha_inicio,
                o.fecha_fin,
                o.plazo
            FROM public."Aplicacion_obra" o
        """)
        hechos_obras = cursor.fetchall()

        # Iniciamos una transacción
        cursor.execute('BEGIN;')
        # Truncar las tablas de hechos antes de insertar nuevos datos
        cursor.execute("TRUNCATE TABLE datamart.hecho_estado_ingeniero;")
        cursor.execute("TRUNCATE TABLE datamart.hecho_certificado;")
        cursor.execute("TRUNCATE TABLE datamart.hecho_presupuesto;")
        cursor.execute("TRUNCATE TABLE datamart.hechomaterialporobra;")
        cursor.execute("TRUNCATE TABLE datamart.hechocronograma;")
        cursor.execute("TRUNCATE TABLE datamart.hecho_obra;")

        # Inserciones para HechoEstadoIngeniero
        for hecho in hechos_estado_ingeniero:
            ingeniero_id, nombre_ingeniero, cantidad_obras_ejecucion, cantidad_presupuestos_elaboracion = hecho

            cursor.execute("""
                        INSERT INTO datamart.hecho_estado_ingeniero (
                            ingeniero_id, nombre_ingeniero, cantidad_obras_ejecucion, cantidad_presupuestos_elaboracion
                        ) 
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (ingeniero_id) DO UPDATE SET
                            nombre_ingeniero = EXCLUDED.nombre_ingeniero,
                            cantidad_obras_ejecucion = EXCLUDED.cantidad_obras_ejecucion,
                            cantidad_presupuestos_elaboracion = EXCLUDED.cantidad_presupuestos_elaboracion;
                    """, [
                ingeniero_id, nombre_ingeniero, cantidad_obras_ejecucion, cantidad_presupuestos_elaboracion
            ])

        # Inserciones para HechoCertificado
        for hecho in hechos_certificado:
            id, ingeniero_id, presupuesto_id, estado, fecha_creacion, fecha_envio, fecha_pago, iva, monto_total, subtotal = hecho

            cursor.execute("""
                    INSERT INTO datamart.hecho_certificado (
                        id, ingeniero_id, presupuesto_id, estado, fecha_creacion, 
                        fecha_envio, fecha_pago, iva, monto_total, subtotal
                    ) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                """, [
                id, ingeniero_id, presupuesto_id, estado, fecha_creacion,
                fecha_envio, fecha_pago, iva, monto_total, subtotal
            ])
        # Inserciones para HechoPresupuesto
        for hecho in hechos_presupuesto:
            id, monto_total, estado, encargado_id, proyecto_id, anticipo, monto_anticipo, fecha_pago_anticipo, comprobante_anticipo, iva, subtotal = hecho

            cursor.execute("""
                        INSERT INTO datamart.hecho_presupuesto (
                            id, monto_total, estado, encargado_id, proyecto_id, anticipo,
                            monto_anticipo, fecha_pago_anticipo, comprobante_anticipo,
                            iva, subtotal
                        ) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """, [
                id, monto_total, estado, encargado_id, proyecto_id, anticipo,
                monto_anticipo, fecha_pago_anticipo, comprobante_anticipo,
                iva, subtotal
            ])
        # Inserciones para HechoMaterialPorObra
        for hecho in hechos_materiales:
            obra_id, material_id, cantidad_total, obra_nombre, material_nombre = hecho

            cursor.execute("""
                        INSERT INTO datamart.hechomaterialporobra (
                            obra_id, material_id, cantidad_total, obra_nombre, material_nombre
                        ) 
                        VALUES (%s, %s, %s, %s, %s);
                    """, [
                obra_id, material_id, cantidad_total, obra_nombre, material_nombre
            ])
        # Inserciones para HechoCronograma
        for hecho in hechos_cronograma:
            cronograma_id, detalle_id, fecha_programada, fecha_culminacion = hecho
            cursor.execute("""
                INSERT INTO datamart.hechocronograma (
                    cronograma_id, detalle_id, fecha_programada, fecha_culminacion
                ) 
                VALUES (%s, %s, %s, %s);
            """, [
                cronograma_id, detalle_id, fecha_programada, fecha_culminacion
            ])
        # Inserciones para HechoObra
        for hecho in hechos_obras:
            id_obra, id_encargado, estado, fecha_inicio, fecha_fin, plazo = hecho
            cursor.execute("""
                INSERT INTO datamart.hecho_obra (
                    id_obra, id_encargado, estado, fecha_inicio, fecha_fin, plazo
                ) 
                VALUES (%s, %s, %s, %s, %s, %s);
            """, [
                id_obra, id_encargado, estado, fecha_inicio, fecha_fin, plazo
            ])
        # Finalizamos la transacción
        cursor.execute('COMMIT;')


def inicio(request):
    verificar_obras_agendadas()
    cargar_distritos()

    if request.user.groups.filter(name='GERENTE').exists():
        actualizar_dimensiones()
        actualizar_hechos()
        ingenieros = HechoEstadoIngeniero.objects.all()
        graph_data_estado_ingenieros = plot_estado_ingenieros_json(ingenieros)
        graph_data_certificados_pendientes = plot_certificados_pendientes()
        obras_activas = Obra.objects.filter(estado='E').select_related('proyecto', 'encargado', 'proyecto__cliente')
        clientes, ingenieros_encargados = get_ingenieros_clientes_obras_activas()
        # Paginación
        paginator = Paginator(obras_activas, 5)  # Mostrar 5 obras por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {'graph_data_estado_ingenieros': json.dumps(graph_data_estado_ingenieros),
                   'graph_data_certificados_pendientes': json.dumps(graph_data_certificados_pendientes),
                   'clientes': clientes,
                   'ingenieros_encargados': ingenieros_encargados,
                   'page_obj': page_obj,
                   }
        return render(request, 'Inicios/inicio.html', context)

    elif request.user.groups.filter(name='ADMINISTRADOR').exists():
        return render(request, 'Inicios/inicio_adm.html')

    elif request.user.groups.filter(name='INGENIERO').exists():
        return redirect(inicio_ingenieros)

    elif request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        return redirect(inicio_deposito)
    else:
        # Redirigir a una página de error o una página predeterminada
        return redirect(inicio_deposito)


def inicio_deposito(request):
    nombre = request.user.first_name

    # Contar los pedidos pendientes
    pedidos_pendientes = Pedido.objects.filter(estado='P').count()

    # Contar las devoluciones pendientes
    devoluciones_pendientes = Devolucion.objects.filter(estado='P').count()

    # Contar los pedidos de compras pendientes
    pedidos_compras_pendientes = PedidoCompra.objects.filter(estado='P').count()

    context = {
        'nombre': nombre,
        'pedidos_pendientes': pedidos_pendientes,
        'devoluciones_pendientes': devoluciones_pendientes,
        'pedidos_compras_pendientes': pedidos_compras_pendientes,
    }

    return render(request, 'Inicios/inicio_deposito.html', context)


def inicio_ingenieros(request):
    nombre = request.user.first_name
    return render(request, 'Inicios/inicio_ingenieros.html', {'nombre': nombre})


def inicio_adm(request):
    nombre = request.user.first_name
    return render(request, 'Inicios/inicio_adm.html', {'nombre': nombre})


# VISTAS PARA USUARIOS


def cargar_distritos():
    # Verificar si ya hay 262 registros en el modelo Ciudad
    if Ciudad.objects.count() != 262:
        file_path = settings.BASE_DIR / "Aplicacion/static/distritos.geojson"
        with open(file_path, 'r') as file:
            data = json.load(file)

            # Obtener una lista de nombres de ciudades y ordenarla alfabéticamente
            nombres_ciudades = [feature['properties']['DIST_DESC_'] for feature in data['features']]

            # Reemplazar "Ñ" por "N" en cada nombre de ciudad
            nombres_ciudades = [nombre.replace('Ñ', 'N') for nombre in nombres_ciudades]
            nombres_ciudades_ordenados = sorted(nombres_ciudades)

            for nombre in nombres_ciudades_ordenados:
                Ciudad.objects.get_or_create(nombre=nombre)


def loguear_usuario(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            try:
                perfil = Personal.objects.get(user=user)
            except:
                return redirect('inicio')
            if perfil.must_change_password:
                return redirect('cambiar_password')
            else:
                return redirect('inicio')
        else:
            error = 'Nombre de usuario o contraseña incorrectos.'
    else:
        error = ''
    return render(request, 'ABM/usuarios/login.html', {'error': error, 'login_page': True} )


@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Importante para mantener la sesión del usuario
            messages.success(request, 'Tu contraseña ha sido actualizada exitosamente.')
            return redirect('inicio')  # Redirige a la URL donde el usuario vea un mensaje de éxito
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'ABM/usuarios/cambiar_password.html', {'form': form})


def salir_usuario(request):
    logout(request)
    return redirect('login')


#                   VISTAS PARA CLIENTE


def registrar_cliente(request):
    if request.method == 'POST':
        # Extraer datos directamente del POST
        nombre = request.POST.get('nombre')
        ruc = request.POST.get('ruc')
        email = request.POST.get('email')
        tipo_persona = request.POST.get('tipo_persona')
        direccion = request.POST.get('direccion')
        id_ciudad = request.POST.get('ciudad')
        ciudad = Ciudad.objects.get(id=id_ciudad)
        telefono = request.POST.get('telefono')
        observaciones = request.POST.get('observaciones')

        # Crear el objeto cliente
        cliente = Cliente(
            nombre=nombre,
            ciudad=ciudad,
            ruc=ruc,
            email=email,
            tipo_persona=tipo_persona,
            direccion=direccion
        )
        cliente.save()
        return redirect('ver_clientes')

    rucs_actuales = list(Cliente.objects.values_list('ruc', flat=True))
    rucs_json = json.dumps(rucs_actuales)
    emails = list(Cliente.objects.values_list('email', flat=True))
    emails_json = json.dumps(emails)

    ciudades = Ciudad.objects.all()  # Obtiene todas las ciudades

    return render(request, 'ABM/clientes/registro_cliente.html',
                  {'rucs_json': rucs_json, 'emails_json': emails_json, 'ciudades': ciudades})


def get_cliente_data(request, cliente_id):
    cliente = Cliente.objects.get(pk=cliente_id)
    data = {
        "nombre": cliente.nombre,
        "ruc": cliente.ruc,
        "email": cliente.email,
        'tipo_persona': cliente.tipo_persona,
        'direccion': cliente.direccion,
        'telefono': cliente.telefono,
        'observaciones': cliente.observaciones
    }

    return JsonResponse(data)


def get_contactos_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    contactos = cliente.contacto_set.all()

    data_contactos = [{"id": contacto.id, "nombre": contacto.nombre, "numero": contacto.numero} for contacto in
                      contactos]
    return JsonResponse({"contactos": data_contactos})


def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo_persona = request.POST.get('tipo_persona')
        ruc = request.POST.get('ruc')
        email = request.POST.get('email')
        direccion = request.POST.get('direccion', '')
        nombre_ciudad = request.POST.get('ciudad', '')

        # Si deseas realizar algunas validaciones básicas, puedes hacerlo aquí
        # Por ejemplo:
        if not nombre or not tipo_persona or not ruc or not email:
            return JsonResponse({"status": "error", "message": "Faltan campos requeridos."}, status=400)

        try:
            if nombre_ciudad:
                ciudad = Ciudad.objects.get(nombre=nombre_ciudad)
                cliente.ciudad = ciudad
            cliente.nombre = nombre
            cliente.tipo_persona = tipo_persona
            cliente.ruc = ruc
            cliente.email = email
            if direccion:
                cliente.direccion = direccion
            cliente.save()

            return JsonResponse({"status": "success"})

        except Ciudad.DoesNotExist:
            return JsonResponse({"status": "error", "message": "La ciudad proporcionada no existe."}, status=400)
        except Exception as e:
            # En caso de cualquier otro error
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return render(request, 'ABM/clientes/editar_cliente.html')


def ver_clientes(request):
    form_buscar = BuscadorClienteForm()
    rucs_actuales = list(Cliente.objects.values_list('ruc', flat=True))
    rucs_json = json.dumps(rucs_actuales)
    emails_actuales = list(Cliente.objects.values_list('email', flat=True))
    emails_json = json.dumps(emails_actuales)
    ciudades = Ciudad.objects.all()

    # Inicializa la consulta
    clientes = Cliente.objects.all()

    # Manejar filtro por ciudad
    ciudad_id = request.GET.get('ciudad')
    if ciudad_id:
        clientes = clientes.filter(ciudad_id=ciudad_id)

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        cliente = get_object_or_404(Cliente, pk=cliente_id)

        proyectos_asociados = Proyecto.objects.filter(cliente=cliente)

        if proyectos_asociados.exists():
            messages.error(request, 'El cliente tiene proyectos activos y no puede ser eliminado')
        else:
            cliente.delete()
            messages.success(request, 'Cliente eliminado exitosamente')
            return redirect('ver_clientes')

    elif request.method == 'GET':
        form_buscar = BuscadorClienteForm(request.GET)
        if form_buscar.is_valid():
            termino_busqueda = form_buscar.cleaned_data['termino_busqueda']
            if termino_busqueda:
                clientes = clientes.filter(Q(ruc__icontains=termino_busqueda) | Q(nombre__icontains=termino_busqueda))

    # Configuración de la paginación después de aplicar los filtros
    paginator = Paginator(clientes, 10)  # Muestra 10 clientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'ABM/clientes/ver_clientes.html', {
        'clientes': clientes,
        'form_buscar': form_buscar,
        'page_obj': page_obj,
        'rucs_json': rucs_json,
        'emails_json': emails_json,
        'ciudades': ciudades,
        'ciudad_seleccionada': ciudad_id,
    })


def buscar_clientes(request):
    query = request.GET.get('q', '')
    clientes = Cliente.objects.filter(
        Q(nombre__icontains=query) |
        Q(ruc__icontains=query)
    )
    data = [{'id': cliente.id, 'nombre': cliente.nombre, 'ruc': cliente.ruc} for cliente in clientes]
    return JsonResponse(data, safe=False)


def ver_cliente(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        cliente = get_object_or_404(Cliente, pk=cliente_id)

        proyectos_asociados = Proyecto.objects.filter(cliente=cliente)

        if proyectos_asociados.exists():
            messages.error(request, 'El cliente tiene proyectos activos y no puede ser eliminado')
        else:
            cliente.delete()
            messages.success(request, 'Cliente eliminado exitosamente')
            return redirect('ver_clientes')

    # Determinar la plantilla base según el grupo del usuario
    if request.user.groups.filter(name="GERENTE").exists():
        base_template = "base_gerente.html"
    elif request.user.groups.filter(name="ADMINISTRADOR").exists():
        base_template = "base_adm.html"
    else:
        base_template = "base_generic.html"  # Asegúrate de que este sea un nombre de plantilla válida

    context = {
        'cliente': cliente,
        'base_template': base_template,
    }

    return render(request, 'ABM/clientes/ver_cliente.html', context)


def agregar_contacto(request, tipo, id):
    if request.method == 'POST':
        data = json.loads(request.body)
        nombre = data.get('nombre')
        numero = data.get('numero')
        contacto = Contacto(nombre=nombre, numero=numero)
        if tipo == 'cliente':
            contacto.cliente = Cliente.objects.get(pk=id)
        elif tipo == 'proveedor':
            contacto.proveedor = Proveedor.objects.get(pk=id)
        else:
            return JsonResponse({'status': 'error', 'message': 'Tipo no válido'})
        contacto.save()

        return JsonResponse({'status': 'success', 'message': 'Contacto agregado correctamente'})


def eliminar_contacto(request, contacto_id):
    if request.method == 'POST':
        contacto = Contacto.objects.get(pk=contacto_id)
        contacto.delete()

        return JsonResponse({'status': 'success', 'message': 'Contacto eliminado correctamente'})


#                   VISTAS PARA INGENIERO

def registrar_ingeniero(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            nombre = form.cleaned_data['first_name']
            apellido = form.cleaned_data['last_name']
            telefono = form.cleaned_data['telefono']
            direccion = form.cleaned_data['direccion']
            password = '12345'  # Obtener la contraseña ingresada

            # Resto del código para validar y crear el usuario

            # Si no hay errores adicionales, creamos el usuario y establecemos la contraseña
            if not form.errors:
                usuario = User.objects.create_user(username=username, first_name=nombre, last_name=apellido,
                                                   email=email)
                usuario.set_password(password)  # Establece la contraseña ingresada por el usuario
                usuario.save()
                grupo = Group.objects.get(name='INGENIERO')
                grupo.user_set.add(usuario)

                return redirect('ver_ingenieros')
        else:
            return redirect('ver_ingenieros')

    else:
        form = CustomUserCreationForm()
        no_obligatorios = ['Teléfono', 'Dirección']
        return render(request, 'ABM/ingenieros/registrar_ingeniero.html',
                      {'form': form, 'no_obligatorios': no_obligatorios})


def ver_ingenieros(request):
    group_ingeniero = Group.objects.get(name='INGENIERO')
    usuarios_ingenieros = group_ingeniero.user_set.all()
    # Configuración de la paginación
    paginator = Paginator(usuarios_ingenieros, 10)  # Muestra 10 ingenieros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'ABM/ingenieros/ver_ingenieros.html', {'page_obj': page_obj})


def editar_ingeniero(request, pk):
    ingeniero = get_object_or_404(User, pk=pk)

    # Comprobamos si es una solicitud POST
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=ingeniero)

        if form.is_valid():
            nombre = request.POST['first_name']
            apellido = request.POST['last_name']
            email = form.cleaned_data['email']

            errors = {}

            if not re.match(r'^[A-Za-z]+$', nombre):
                errors["first_name"] = "Nombre inválido. Solo se permiten letras."
            if not re.match(r'^[A-Za-z]+$', apellido):
                errors["last_name"] = "Apellido inválido. Solo se permiten letras."
            if User.objects.filter(email=email).exclude(pk=pk).exists():
                return JsonResponse({"status": "error", "errors": {"email": ["El email ya está registrado."]}})
            if errors.keys():
                return JsonResponse({"status": "error", "errors": errors})
            else:
                form.save()
                return JsonResponse({"status": "success"})
        else:
            print(form.errors)
            return JsonResponse({"status": "error", "errors": form.errors})

    return JsonResponse({"status": "error", "message": "Metodo no permitido"})


def get_ingeniero_data(request, ingeniero_id):
    ingeniero = get_object_or_404(User, pk=ingeniero_id)
    data = {
        "first_name": ingeniero.first_name,
        "last_name": ingeniero.last_name,
        "email": ingeniero.email,
        # ... otros campos ...
    }
    return JsonResponse(data)


def eliminar_ingeniero(request, pk):
    ingeniero = get_object_or_404(User, pk=pk)

    # Verificamos si el ingeniero tiene obras o presupuestos asociados que no estén en los estados mencionados.
    obras_asociadas = Obra.objects.filter(encargado=ingeniero).exclude(estado='F')
    presupuestos_asociados = Presupuesto.objects.filter(encargado=ingeniero).exclude(estado='A')
    # Si tiene recursos asociados, obtenemos una lista de todos los ingenieros disponibles para la reasignación
    if obras_asociadas.exists() or presupuestos_asociados.exists():
        group_ingeniero = Group.objects.get(name='INGENIERO')
        otros_ingenieros = group_ingeniero.user_set.exclude(id=ingeniero.id)
    else:
        otros_ingenieros = []

    if request.method == 'POST' and 'confirmar' in request.POST:
        reasignaciones_completas = True
        # Reasignamos cada recurso al ingeniero seleccionado en la lista desplegable
        if obras_asociadas:
            for obra in obras_asociadas:
                nuevo_ingeniero_id = request.POST.get(f"reassign_obra_{obra.id}")
                if nuevo_ingeniero_id:
                    obra_original = get_object_or_404(Obra, pk=obra.pk)
                    obra_original.encargado = User.objects.get(pk=nuevo_ingeniero_id)
                    obra_original.save()
                else:
                    reasignaciones_completas = False
        if presupuestos_asociados:
            for presupuesto in presupuestos_asociados:
                nuevo_ingeniero_id = request.POST.get(f"reassign_presupuesto_{presupuesto.id}")
                if nuevo_ingeniero_id:
                    presupuesto_original = get_object_or_404(Presupuesto, pk=presupuesto.pk)
                    presupuesto_original.encargado = User.objects.get(pk=nuevo_ingeniero_id)
                    presupuesto_original.save()
                else:
                    reasignaciones_completas = False
        # Ahora podemos eliminar al ingeniero original
        if reasignaciones_completas:
            ingeniero.delete()
            return redirect('ver_ingenieros')
        else:
            messages.error(request,
                           'No se pudo eliminar el ingeniero porque no se asignaron otros ingenieros a los recursos asociados.')
    context = {
        'ingeniero': ingeniero,
        'obras_asociadas': obras_asociadas,
        'presupuestos_asociados': presupuestos_asociados,
        'otros_ingenieros': otros_ingenieros
    }
    return render(request, 'ABM/ingenieros/eliminar_ingeniero.html', context)


def buscar_ingenieros(request):
    query = request.GET.get('q', '')
    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query)
    )

    data = [{'id': ingeniero.id, 'username': ingeniero.username, 'nombre': ingeniero.first_name,
             'apellido': ingeniero.last_name} for ingeniero in ingenieros]
    return JsonResponse(data, safe=False)


def ver_ingeniero(request, id_ingeniero):
    ingeniero = User.objects.get(pk=id_ingeniero)
    return render(request, 'ABM/ingenieros/ver_ingeniero.html', {'ingeniero': ingeniero})


#                   VISTAS PARA PROVEEDOR


def registrar_proveedor(request):
    if request.method == 'POST':
        # Extraer datos directamente del POST
        nombre = request.POST.get('nombre')
        ruc = request.POST.get('ruc')
        email = request.POST.get('email')
        direccion = request.POST.get('direccion')
        id_ciudad = request.POST.get('ciudad')
        ciudad = Ciudad.objects.get(id=id_ciudad)
        telefono = request.POST.get('telefono')
        observaciones = request.POST.get('observaciones')
        pagina_web = request.POST.get('pagina_web')

        # Crear el objeto cliente
        proveedor = Proveedor(
            nombre=nombre,
            ciudad=ciudad,
            ruc=ruc,
            email=email,
            telefono=telefono,
            direccion=direccion,
            observaciones=observaciones,
            pagina_web=pagina_web
        )
        proveedor.save()
        return redirect(ver_proveedores)
    else:
        rucs_actuales = list(Proveedor.objects.values_list('ruc', flat=True))
        rucs_json = json.dumps(rucs_actuales)
        emails = list(Proveedor.objects.values_list('email', flat=True))
        emails_json = json.dumps(emails)
        ciudades = Ciudad.objects.all()  # Obtiene todas las ciudades
        return render(request, 'ABM/proveedores/registrar_proveedor.html',
                      {'rucs_json': rucs_json, 'emails_json': emails_json, 'ciudades': ciudades})


def ver_pedidos_deposito(request):
    # Obtener filtros de la solicitud GET
    estado = request.GET.get('estado')
    obra_id = request.GET.get('obra')
    encargado_id = request.GET.get('encargado')
    # Filtrar pedidos en base a los filtros proporcionados
    pedidos = Pedido.objects.all().order_by('-estado', '-fecha_solicitud')
    if estado:
        pedidos = pedidos.filter(estado=estado)
    if obra_id:
        pedidos = pedidos.filter(obra_id=obra_id)
    if encargado_id:
        pedidos = pedidos.filter(obra__encargado_id=encargado_id)

    # Obtener listas para los filtros
    obras = Obra.objects.all()
    encargados = User.objects.filter(groups__name='INGENIERO')

    context = {
        'pedidos': pedidos,
        'obras': obras,
        'encargados': encargados,
        'estado_seleccionado': estado,
        'obra_seleccionada': obra_id,
        'encargado_seleccionado': encargado_id,
    }
    return render(request, 'pantallas_deposito/ver_pedidos_dep.html', context)


def ver_proveedores(request):
    proveedores = Proveedor.objects.all()
    form_buscar = BuscadorProveedorForm()
    rucs_actuales = list(Proveedor.objects.values_list('ruc', flat=True))
    rucs_json = json.dumps(rucs_actuales)
    emails_actuales = list(Proveedor.objects.values_list('email', flat=True))
    emails_json = json.dumps(emails_actuales)
    ciudades = Ciudad.objects.all()

    # Obtener la ciudad seleccionada de la URL
    ciudad_id = request.GET.get('ciudad')

    # Filtrar proveedores por ciudad si se ha seleccionado una ciudad
    if ciudad_id:
        proveedores = proveedores.filter(ciudad_id=ciudad_id)

    # Configuración de la paginación
    paginator = Paginator(proveedores, 10)  # Muestra 10 proveedores por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'GET':
        form_buscar = BuscadorProveedorForm(request.GET)
        if form_buscar.is_valid():
            termino_busqueda = form_buscar.cleaned_data['termino_busqueda']
            if termino_busqueda:
                if termino_busqueda.isdigit():
                    proveedores = proveedores.filter(ruc__icontains=termino_busqueda)
                else:
                    proveedores = proveedores.filter(nombre__icontains=termino_busqueda)

    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_proveedores_dep.html'
    else:
        template_name = 'ABM/proveedores/ver_proveedores.html'

    return render(request, template_name, {
        'proveedores': proveedores,
        'form_buscar': form_buscar,
        'page_obj': page_obj,
        'rucs_json': rucs_json,
        'emails_json': emails_json,
        'ciudades': ciudades,
        'ciudad_seleccionada': ciudad_id,  # Pasar la ciudad seleccionada al template
    })


def get_proveedor_data(request, proveedor_id):
    proveedor = Proveedor.objects.get(pk=proveedor_id)
    data = {
        "nombre": proveedor.nombre,
        "ruc": proveedor.ruc,
        "email": proveedor.email,
        'ciudad': proveedor.ciudad.nombre,
        'direccion': proveedor.direccion,
        'pagina_web': proveedor.pagina_web,
        'telefono': proveedor.telefono,
        'observaciones': proveedor.observaciones
        # ... otros campos ...
    }
    return JsonResponse(data)


def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)

    # Verificamos si el proveedor tiene materiales asociados
    materiales_asociados = Material.objects.filter(id_proveedor=proveedor.pk)

    # Si tiene materiales asociados, obtenemos una lista de todos los proveedores disponibles para la reasignación
    if materiales_asociados.exists():
        otros_proveedores = Proveedor.objects.exclude(pk=proveedor.pk).all()
    else:
        otros_proveedores = []

    if request.method == 'POST' and 'confirmar' in request.POST:
        reasignaciones_completas = True

        # Reasignamos cada material al proveedor seleccionado en la lista desplegable
        for material in materiales_asociados:
            nuevo_proveedor_id = request.POST.get(f"reassign_material_{material.id}")
            if nuevo_proveedor_id:
                material_original = get_object_or_404(Material, pk=material.pk)
                material_original.id_proveedor = Proveedor.objects.get(pk=nuevo_proveedor_id)
                material_original.save()
            else:
                reasignaciones_completas = False

        # Si todas las reasignaciones están completas, eliminamos al proveedor
        if reasignaciones_completas:
            proveedor.delete()
            return redirect('ver_proveedores')
        else:
            # Añadimos un mensaje indicando que no se pudo eliminar al proveedor
            messages.error(request,
                           'El proveedor tiene materiales asociados, reasignelos a otro proveedor o elimine el material desde la pantalla de Materiales.')

    context = {
        'proveedor': proveedor,
        'materiales_asociados': materiales_asociados,
        'otros_proveedores': otros_proveedores
    }
    return render(request, 'ABM/proveedores/eliminar_proveedor.html', context)


def get_contactos_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    contactos = proveedor.contacto_set.all()

    data_contactos = [{"id": contacto.id, "nombre": contacto.nombre, "numero": contacto.numero} for contacto in
                      contactos]
    return JsonResponse({"contactos": data_contactos})


def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)

    if request.method == 'POST':
        # Actualizamos los atributos del proveedor con los datos recibidos
        proveedor.nombre = request.POST.get('nombre', proveedor.nombre)
        proveedor.ruc = request.POST.get('ruc', proveedor.ruc)
        proveedor.email = request.POST.get('email', proveedor.email)
        proveedor.direccion = request.POST.get('direccion', proveedor.direccion)
        nombre_ciudad = request.POST.get('ciudad')
        proveedor.ciudad = Ciudad.objects.get(nombre=nombre_ciudad)
        proveedor.telefono = request.POST.get('telefono', proveedor.telefono)
        proveedor.pagina_web = request.POST.get('pagina_web', proveedor.pagina_web)
        proveedor.observaciones = request.POST.get('observaciones', proveedor.observaciones)
        # Guardamos el objeto proveedor
        try:
            proveedor.full_clean()  # Valida el objeto antes de guardarlo
            proveedor.save()
            return JsonResponse({"status": "success", "message": "Proveedor actualizado con éxito"})
        except ValidationError as e:
            # En caso de que el objeto no sea válido, devolvemos los errores en formato JSON
            return JsonResponse({"status": "error", "errors": e.message_dict}, status=400)

    return render(request, 'ABM/proveedores/editar_proveedor.html')


def buscar_proveedores(request):
    query = request.GET.get('q', '')
    proveedores = Proveedor.objects.filter(
        Q(nombre__icontains=query) |
        Q(ruc__icontains=query)
    )
    data = [{'id': proveedor.id, 'nombre': proveedor.nombre, 'ruc': proveedor.ruc} for proveedor in proveedores]
    return JsonResponse(data, safe=False)


def ver_proveedor(request, id_proveedor):
    proveedor = Proveedor.objects.get(pk=id_proveedor)
    return render(request, 'ABM/proveedores/ver_proveedor.html', {'proveedor': proveedor})


#                   VISTAS PARA MATERIAL


def registrar_material(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('ver_materiales')
    else:
        form = MaterialForm()
    return render(request, 'ABM/materiales/registrar_material.html', {'form': form})


def ver_materiales(request):
    materiales = Material.objects.all()
    form_buscar = BuscadorMaterialForm()

    # Configuración de la paginación
    paginator = Paginator(materiales, 10)  # Muestra 10 materiales por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lógica para eliminar el material
    material_id = request.GET.get('eliminar_material_id')
    if material_id:
        material = get_object_or_404(Material, id=material_id)

        # Verificamos si el material está siendo utilizado en pedidos
        pedidos_asociados = Pedido.objects.filter(materiales=material)

        # Verificamos si las obras asociadas a esos pedidos están activas
        obras_activas_asociadas = Obra.objects.filter(pedido__in=pedidos_asociados, estado='E')

        if obras_activas_asociadas.exists():
            # Si hay obras activas que hacen referencia al material, mostramos un mensaje de error
            messages.error(request, 'No se puede eliminar el material porque está siendo utilizado en obras activas.')
        else:
            material.delete()
            messages.success(request, 'Material eliminado exitosamente.')

    if request.method == 'GET':
        form_buscar = BuscadorMaterialForm(request.GET)
        if form_buscar.is_valid():
            termino_busqueda = form_buscar.cleaned_data['termino_busqueda']
            if termino_busqueda:
                materiales = materiales.filter(nombre__icontains=termino_busqueda)

    return render(request, 'ABM/materiales/ver_materiales.html',
                  {'materiales': materiales, 'form_buscar': form_buscar, 'page_obj': page_obj})


def editar_material(request, pk):
    material = get_object_or_404(Material, id=pk)

    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            form.save()
            return redirect('ver_materiales')
    else:
        form = MaterialForm(instance=material)

    return render(request, 'ABM/materiales/editar_material.html', {'form': form})


def buscar_materiales(request):
    query = request.GET.get('q', '')
    materiales = Material.objects.filter(
        Q(nombre__icontains=query)
    )
    data = [{'id': material.id, 'nombre': material.nombre} for material in materiales]
    return JsonResponse(data, safe=False)


def ver_material(request, id_material):
    material = Material.objects.get(pk=id_material)
    return render(request, 'ABM/materiales/ver_material.html', {'material': material})


#                   VISTAS PARA PROYECTOS


def registrar_proyecto(request):
    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.all()
    clientes = Cliente.objects.all()
    ciudades = Ciudad.objects.all()  # Obtiene todas las ciudades

    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            presupuesto = Presupuesto(encargado=form.cleaned_data['encargado'])
            proyecto.cliente = form.cleaned_data['cliente']
            proyecto.nombre = request.POST.get('nombre')
            ciudad_id = request.POST.get('ciudad')
            ciudad = Ciudad.objects.get(id=ciudad_id)
            proyecto.ciudad = ciudad
            obra = Obra()

            # Primero guardamos el proyecto para obtener su ID
            proyecto.save()

            # Luego asignamos el proyecto a las instancias de Presupuesto y Obra
            presupuesto.proyecto = proyecto
            presupuesto.save()
            obra.proyecto = proyecto
            obra.save()

            # Ahora asignamos las instancias de Presupuesto y Obra al Proyecto
            proyecto.presupuesto = presupuesto
            proyecto.obra = obra
            proyecto.save()

            return redirect('ver_proyectos')
    else:
        form = ProyectoForm()
    return render(request, 'ABM/proyectos/registrar_proyecto.html',
                  {'form': form, 'ingenieros': ingenieros, 'clientes': clientes, 'ciudades': ciudades})


def ver_proyectos(request):
    proyectos = Proyecto.objects.all()
    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'ABM/proyectos/ver_proyectos.html', {'proyectos': proyectos, 'page_obj': page_obj})


def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)

    if request.method == 'POST' and 'confirmar' in request.POST:
        proyecto.delete()
        return redirect('ver_proyectos')

    return render(request, 'ABM/proyectos/eliminar_proyecto.html', {'proyecto': proyecto})


def modificar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.all()
    clientes = Cliente.objects.all()
    if request.method == 'POST':
        cliente = request.POST['cliente']
        encargado_obra = request.POST['encargado_obra']
        encargado_presupuesto = request.POST['encargado_presupuesto']
        nombre_proyecto = request.POST['nombre']
        proyecto.cliente = Cliente.objects.get(id=cliente)
        proyecto.nombre = nombre_proyecto
        if encargado_presupuesto:
            proyecto.presupuesto.encargado = User.objects.get(id=encargado_presupuesto)
            proyecto.presupuesto.save()
        if encargado_obra != '':
            proyecto.obra.encargado = User.objects.get(id=encargado_obra)
            proyecto.obra.save()
        proyecto.save()
        return redirect('ver_proyectos')
    return render(request, 'ABM/proyectos/modificar_proyecto.html',
                  {'proyecto': proyecto, 'clientes': clientes, 'ingenieros': ingenieros})


def buscar_proyectos(request):
    query = request.GET.get('q', '')
    proyectos = Proyecto.objects.filter(
        Q(nombre__icontains=query)
    )
    data = [{'id': proyecto.id, 'nombre': proyecto.nombre} for proyecto in proyectos]
    return JsonResponse(data, safe=False)


def ver_proyecto(request, id_proyecto):
    proyecto = Proyecto.objects.get(pk=id_proyecto)
    return render(request, 'ABM/proyectos/ver_proyecto.html', {'proyecto': proyecto})


def ver_presupuestos(request):
    presupuestos = Presupuesto.objects.filter(encargado=request.user)

    return render(request, 'pantallas_ing/ver_presupuestos.html', {'presupuestos': presupuestos})


def ver_obras(request):
    obras = Obra.objects.filter(encargado=request.user)

    # Iterar sobre las obras y agregar parámetros adicionales
    obras_con_parametros = []
    for obra in obras:
        cronograma = Cronograma.objects.get(archivo_presupuesto__presupuesto__proyecto=obra.proyecto)
        # Obtener el cronograma asociado a la obra
        # Verificar si existe un cronograma y si al menos un detalle tiene fecha
        if cronograma and cronograma.detalles_cronograma.filter(fecha_programada__isnull=False).exists():
            cronograma_existente = True
        else:
            cronograma_existente = False
        print(cronograma_existente)
        obra_parametros = {
            'obra': obra,
            'cronograma_existe': cronograma_existente,
            'cronograma': cronograma
        }
        obras_con_parametros.append(obra_parametros)

    return render(request, 'pantallas_ing/ver_obras.html', {'obras': obras_con_parametros})


def pedido_materiales(request):
    obras = Obra.objects.filter(encargado=request.user)
    search_query = request.GET.get('search', '')

    # Filtrar los materiales por nombre o marca según la búsqueda
    if search_query:
        materiales = Material.objects.filter(
            Q(nombre__icontains=search_query) |
            Q(marca__icontains=search_query)
        )
    else:
        # Si no hay búsqueda, mostrar todos los materiales
        materiales = Material.objects.all()

    if request.method == 'POST':
        materiales_pedido = []
        obra = get_object_or_404(Obra, id=request.POST.get('obra'))
        for key, value in request.POST.items():
            if key.startswith('material_'):
                material_id = int(key.split('_')[1])
                material = get_object_or_404(Material, id=material_id)
                cantidad = int(value)
                materiales_pedido.append({'material': material, 'cantidad': cantidad})

        # Por ejemplo, puedes crear las instancias MaterialPedido aquí.
        return render(request, 'pantallas_ing/confirmar_pedido.html',
                      {'materiales_pedido': materiales_pedido, 'obra': obra})

    # Paginación de materiales
    paginator = Paginator(materiales, 12)  # Muestra 18 materiales por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Asegúrate de que el contexto incluya 'page_obj' y no 'materiales' directamente
    context = {
        'obras': obras,
        'page_obj': page_obj,
        'search_query': search_query
    }
    return render(request, 'pantallas_ing/pedido_materiales.html', context)


def confirmar_pedido(request):
    if request.method == 'POST':
        # Obtener los materiales seleccionados del formulario
        materiales_pedido = []
        for key, value in request.POST.items():
            if key.startswith('cantidad_') and int(value) > 0:
                material_id = key.split('_')[1]
                material = Material.objects.get(id=material_id)
                cantidad = int(value)
                materiales_pedido.append({'material': material, 'cantidad': cantidad})

        # Crear el objeto Pedido y guardar en la base de datos
        pedido = Pedido.objects.create(
            solicitante=request.user,
            obra=Obra.objects.get(id=request.POST.get('obra')),
            fecha_solicitud=date.today(),
            fecha_entrega=None,
            estado='P'
        )

        # Crear los objetos MaterialPedido y guardar en la base de datos
        for material_pedido in materiales_pedido:
            material = material_pedido['material']
            cantidad = material_pedido['cantidad']
            MaterialPedido.objects.create(pedido=pedido, material=material, cantidad=cantidad)

        # Redirigir a una página de éxito o realizar alguna acción adicional
        return redirect('inicio_ingenieros')  # Reemplaza 'pagina_de_exito' con la URL a

def confirmar_pedido_compra(request):
    if request.method == 'POST':
        # Obtener los materiales seleccionados del formulario
        materiales_pedido = []
        for key, value in request.POST.items():
            if key.startswith('cantidad_') and int(value) > 0:
                material_id = key.split('_')[1]
                material = get_object_or_404(Material, id=material_id)
                cantidad = int(value)
                materiales_pedido.append({'material': material, 'cantidad': cantidad})


        # Crear el objeto PedidoCompra y guardar en la base de datos
        pedido_compra = PedidoCompra.objects.create(
            fecha_solicitud=date.today(),
            fecha_entrega=None,
            estado='P'
        )

        # Crear los objetos MaterialPedidoCompra y guardar en la base de datos
        for material_pedido in materiales_pedido:
            material = material_pedido['material']
            cantidad = material_pedido['cantidad']
            MaterialPedidoCompra.objects.create(pedido_compra=pedido_compra, material=material, cantidad=cantidad)

        # Pasar el pedido y los materiales al contexto
        context = {
            'pedido_compra': pedido_compra,
            'materiales_pedido': materiales_pedido,
        }

        # Redirigir a la vista que muestra los pedidos
        return redirect('ver_pedidos_compras')

    # Si el método no es POST, redirigir a una página de error o a la lista de materiales
    return redirect('ver_pedidos_compras')

@csrf_exempt
def actualizar_pedido_compra(request, pedido_id):
    if request.method == "POST":
        try:
            pedido = get_object_or_404(PedidoCompra, pk=pedido_id)
            estado = request.POST.get('estado')
            if estado:
                pedido.estado = estado

            # Si se ha enviado un archivo de comprobante, lo guardamos
            comprobante = request.FILES.get('comprobante')
            if comprobante:
                pedido.comprobante = comprobante

            pedido.fecha_entrega = date.today()

            pedido.save()

            # Actualizar el stock de los materiales asociados con el pedido
            materiales_pedido = MaterialPedidoCompra.objects.filter(pedido_compra=pedido)
            for material_pedido in materiales_pedido:
                material = material_pedido.material
                material.unidades_stock += material_pedido.cantidad
                material.save()

            return JsonResponse({
                'status': 'success',
                'nuevo_estado': pedido.get_estado_display(),
                'fecha_entrega': pedido.fecha_entrega.strftime('%d/%m/%Y') if pedido.fecha_entrega else None
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

def ver_pedidos(request):
    # Obtener todos los pedidos del usuario actualmente logueado
    pedidos = Pedido.objects.filter(solicitante=request.user).order_by('-estado', 'fecha_solicitud')

    context = {
        'pedidos': pedidos
    }
    return render(request, 'pantallas_ing/ver_pedidos.html', context)


def ver_pedidos_compras(request):
    # Filtrar pedidos pendientes que tengan al menos un material con cantidad mayor a cero
    pedidos = PedidoCompra.objects.all().order_by('estado', '-fecha_solicitud')

    pedidos_con_materiales = []
    for pedido in pedidos:
        # Obtener materiales con cantidad mayor a cero para cada pedido
        materiales = MaterialPedidoCompra.objects.filter(pedido_compra=pedido, cantidad__gt=0).select_related(
            'material')

        # Asegurarse de que solo se incluyan pedidos que realmente tienen materiales
        if materiales.exists():
            pedidos_con_materiales.append({
                'pedido': pedido,
                'materiales': materiales,
            })

    context = {
        'pedidos_con_materiales': pedidos_con_materiales
    }
    return render(request, 'pantallas_deposito/ver_pedidos_compras.html', context)

def ver_pedido_compras(request, pedido_id):
    pedido = get_object_or_404(PedidoCompra, pk=pedido_id)
    materiales = MaterialPedidoCompra.objects.filter(pedido_compra=pedido).select_related('material')

    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_pedido_compras.html'
    else:
        template_name = 'pantallas_adm/ver_compra_adm.html'

    context = {
        'pedido': pedido,
        'materiales': materiales
    }
    return render(request, template_name, context)

def pedido_compra(request):
    # 1. Materiales con stock menor al mínimo
    materiales_bajo_minimo = Material.objects.filter(unidades_stock__lte=F('minimo'))

    # 2. Materiales en pedidos pendientes con cantidad solicitada mayor al stock
    materiales_en_pedidos_pendientes = MaterialPedido.objects.filter(
        pedido__estado='P',
        cantidad__gt=F('material__unidades_stock')
    ).select_related('material', 'pedido')

    # 3. Excluir materiales que ya están en pedidos de compra pendientes
    materiales_en_pedidos_compra_pendientes = MaterialPedidoCompra.objects.filter(
        pedido_compra__estado='P'
    ).values_list('material_id', flat=True).distinct()

    # Crear una lista de materiales faltantes con información adicional
    materiales_con_info = []

    # Agregar materiales con stock menor al mínimo
    for material in materiales_bajo_minimo:
        if material.id not in materiales_en_pedidos_compra_pendientes:
            cantidad_faltante = material.minimo - material.unidades_stock

            # Obtener la cantidad en pedidos pendientes para este material
            cantidad_pedido_pendiente = MaterialPedido.objects.filter(
                pedido__estado='P',
                material=material
            ).aggregate(total_cantidad=Sum('cantidad'))['total_cantidad'] or 0

            # Obtener la cantidad en pedidos de compra pendientes para este material
            cantidad_pedido_compra_pendiente = MaterialPedidoCompra.objects.filter(
                pedido_compra__estado='P',
                material=material
            ).aggregate(total_cantidad=Sum('cantidad'))['total_cantidad'] or 0

            materiales_con_info.append({
                'material': material,
                'cantidad_faltante': cantidad_faltante,
                'cantidad_pedido_pendiente': cantidad_pedido_pendiente,
                'cantidad_pedido_compra_pendiente': cantidad_pedido_compra_pendiente
            })

    # Agregar materiales en pedidos pendientes con cantidad solicitada mayor al stock
    for material_pedido in materiales_en_pedidos_pendientes:
        cantidad_faltante = material_pedido.cantidad - material_pedido.material.unidades_stock
        materiales_con_info.append({
            'material': material_pedido.material,
            'cantidad_faltante': cantidad_faltante,
            'cantidad_pedido_pendiente': material_pedido.cantidad,
            'cantidad_pedido_compra_pendiente': 0  # Si no está en pedidos de compra pendientes
        })

    # Paginación de materiales
    paginator = Paginator(materiales_con_info, 12)  # Muestra 12 materiales por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    return render(request, 'pantallas_deposito/pedido_compra.html', context)

def ver_pedidos_compras_adm(request):
    pedidos = PedidoCompra.objects.all().order_by('estado')
    pedidos_con_materiales = []
    for pedido in pedidos:
        # Obtener materiales con cantidad mayor a cero para cada pedido
        materiales = MaterialPedidoCompra.objects.filter(pedido_compra=pedido, cantidad__gt=0).select_related(
            'material')

        # Asegurarse de que solo se incluyan pedidos que realmente tienen materiales
        if materiales.exists():
            pedidos_con_materiales.append({
                'pedido': pedido,
                'materiales': materiales,
            })
    context = {
        'pedidos_con_materiales': pedidos_con_materiales
    }
    print(pedidos)
    print(pedidos_con_materiales)
    return render(request, 'pantallas_adm/ver_pedidos_compras_adm.html', context)


def ver_pedidos_adm(request, obra_id):
    obra = Obra.objects.get(id=obra_id)
    # Obtener todos los pedidos del usuario actualmente logueado
    pedidos = Pedido.objects.filter(obra=obra)
    # Obtener la suma de la cantidad de cada material en la obra
    materiales_con_cantidad = Material.objects.filter(materialpedido__pedido__obra=obra).annotate(
        total_cantidad=Sum('materialpedido__cantidad')
    )
    context = {
        'obra': obra,
        'pedidos': pedidos,
        'materiales_con_cantidad': materiales_con_cantidad
    }

    return render(request, 'pantallas_adm/ver_pedidos_adm.html', context)

def ver_obras_terminadas(request):
    if request.method == 'POST':
        obra_id = request.POST.get('obra')
        if obra_id:
            return redirect('ver_pedidos_obra', obra_id=obra_id)

    obras_terminadas = Obra.objects.filter(estado='F')
    context = {
        'obras_terminadas': obras_terminadas,
    }
    return render(request, 'pantallas_ing/ver_obras_terminadas.html', context)

def ver_pedidos_obra(request, obra_id):
    obra = Obra.objects.get(pk=obra_id)
    pedidos_entregados = Pedido.objects.filter(obra=obra, estado='E')
    devoluciones = Devolucion.objects.filter(
        pedido__in=pedidos_entregados)  # Obtiene las devoluciones relacionadas con los pedidos entregados

    context = {
        'obra': obra,
        'pedidos': pedidos_entregados,
        'devoluciones': devoluciones,
    }
    return render(request, 'pantallas_ing/ver_pedidos_obra.html', context)

def ver_pedido_a_devolver(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    materiales_pedido = MaterialPedido.objects.filter(pedido=pedido)  # Obtiene los materiales relacionados con el pedido
    return render(request, 'pantallas_ing/ver_pedido_a_devolver.html', {'pedido': pedido, 'materiales_pedido': materiales_pedido})

def confirmar_devolucion(request):
    if request.method == 'POST':
        pedido_id = request.POST.get('pedido_id')
        pedido = Pedido.objects.get(id=pedido_id)

        # Obtener los materiales devueltos del formulario
        materiales_devueltos = []
        for key, value in request.POST.items():
            if key.startswith('cantidad_') and int(value) > 0:
                material_id = key.split('_')[1]
                material = Material.objects.get(id=material_id)
                cantidad = int(value)
                materiales_devueltos.append({'material': material, 'cantidad': cantidad})

        # Crear el objeto Devolucion y guardar en la base de datos
        devolucion = Devolucion.objects.create(
            ingeniero=request.user,
            obra=pedido.obra,
            fecha_solicitud=date.today(),
            fecha_devolucion=None,
            estado='P',
            pedido=pedido
        )

        # Crear los objetos MaterialDevuelto y guardar en la base de datos
        for material_devuelto in materiales_devueltos:
            material = material_devuelto['material']
            cantidad = material_devuelto['cantidad']
            MaterialDevuelto.objects.create(devolucion=devolucion, material=material, cantidad=cantidad)

        # Redirigir a una página de éxito o realizar alguna acción adicional
        return redirect('ver_pedidos_obra', obra_id=pedido.obra.id)
    return redirect('ver_ingenieros')



# vistas para filtros

def obtener_clientes_con_proyectos(request):
    clientes = Proyecto.objects.values_list('cliente__nombre', flat=True).distinct()
    return JsonResponse(list(clientes), safe=False)


def obtener_clientes_con_obras(request):
    obras = Obra.objects.filter(encargado=request.user)
    clientes = obras.values_list('proyecto__cliente__nombre', flat=True).distinct()

    return JsonResponse(list(clientes), safe=False)


def obtener_ingenieros_presupuesto(request):
    ingenieros = Presupuesto.objects.values_list('encargado__username', flat=True).distinct()
    return JsonResponse(list(ingenieros), safe=False)


def obtener_estados_presupuesto(request):
    # Obtener los estados que realmente están en la base de datos
    estados_presentes = Presupuesto.objects.values_list('estado', flat=True).distinct()
    # Convertir esos estados a sus representaciones legibles
    estados = [dict(Presupuesto.ESTADOS)[estado] for estado in estados_presentes]
    return JsonResponse(list(estados), safe=False)


def obtener_estados_obras_ing(request):
    # Obtener los estados que realmente están en la base de datos
    obras = Obra.objects.filter(encargado=request.user)
    estados_presentes = obras.values_list('estado', flat=True).distinct()
    # Convertir esos estados a sus representaciones legibles
    estados = [dict(Obra.ESTADOS)[estado] for estado in estados_presentes]
    return JsonResponse(list(estados), safe=False)


def obtener_ingenieros_obra(request):
    # Excluyendo encargados que son None
    ingenieros = Obra.objects.exclude(encargado__isnull=True).values_list('encargado__username', flat=True).distinct()

    # Convertir a lista y agregar "No asignado"
    ingenieros_list = list(ingenieros)
    ingenieros_list.append("No asignado")
    return JsonResponse(ingenieros_list, safe=False)


def obtener_estados_obra(request):
    # Obtener los estados que realmente están en la base de datos
    estados_presentes = Obra.objects.values_list('estado', flat=True).distinct()
    # Convertir esos estados a sus representaciones legibles
    estados = [dict(Obra.ESTADOS)[estado] for estado in estados_presentes]
    return JsonResponse(list(estados), safe=False)


def obtener_ciudades_con_proyectos(request):
    ciudades_ids = Proyecto.objects.values_list('ciudad', flat=True).distinct()
    nombres_ciudades = [get_object_or_404(Ciudad, id=ciudad_id).nombre for ciudad_id in ciudades_ids]

    # Envía una respuesta JSON con una clave 'data' que contiene la lista de nombres de ciudades
    response_data = {
        'data': nombres_ciudades
    }

    return JsonResponse(response_data)


def ver_proyectos_cliente(request, cliente_nombre):
    proyectos = Proyecto.objects.filter(cliente__nombre=cliente_nombre)
    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos, 'page_obj': page_obj})


def ver_proyectos_encargado_presupuesto(request, ingeniero_user):
    ingeniero = User.objects.filter(username=ingeniero_user)
    proyectos = Proyecto.objects.filter(presupuesto__encargado__in=ingeniero)
    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos, 'page_obj': page_obj})


def ver_proyectos_encargado_obra(request, ingeniero_user):
    if ingeniero_user == "No asignado":
        proyectos = Proyecto.objects.filter(obra__encargado__isnull=True)
    else:
        ingeniero = User.objects.get(username=ingeniero_user)
        proyectos = Proyecto.objects.filter(obra__encargado=ingeniero)

    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos, 'page_obj': page_obj})


def ver_proyectos_estado_presupuesto(request, estado):
    # Convertir el nombre legible del estado a su código correspondiente
    estado_codigo = {value: key for key, value in dict(Presupuesto.ESTADOS).items()}[estado]

    # Filtrar los proyectos basándose en el código del estado
    proyectos = Proyecto.objects.filter(presupuesto__estado=estado_codigo)
    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos, 'page_obj': page_obj})


def ver_proyectos_estado_obra(request, estado):
    print(estado)
    # Convertir el nombre legible del estado a su código correspondiente
    estado_codigo = {value: key for key, value in dict(Obra.ESTADOS).items()}[estado]
    print(estado_codigo)
    # Filtrar los proyectos basándose en el código del estado
    proyectos = Proyecto.objects.filter(obra__estado=estado_codigo)

    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos})


def ver_proyectos_ciudad(request, ciudad):
    proyectos = Proyecto.objects.filter(ciudad=ciudad)
    # Configuración de la paginación
    paginator = Paginator(proyectos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'ABM/proyectos/ver_proyectos_filtrados.html', {'proyectos': proyectos, 'page_obj': page_obj})


def obtener_materiales_marca(request):
    marcas = Material.objects.values_list('marca__nombre', flat=True).distinct().order_by('marca__nombre')
    return JsonResponse(list(marcas), safe=False)


def obtener_materiales_proveedor(request):
    id_proveedores = Material.objects.values_list('id_proveedor', flat=True).distinct()
    lista_proveedores = Proveedor.objects.filter(id__in=id_proveedores)
    proveedores = lista_proveedores.values_list('nombre', flat=True).distinct()
    return JsonResponse(list(proveedores), safe=False)


def obtener_materiales_stock(request):
    cantidades = ['Menos de 10', 'Menos de 50', 'Menos de 100']
    return JsonResponse(list(cantidades), safe=False)


def ver_materiales_marca(request, marca_nombre):
    # Buscar el objeto Marca correspondiente al nombre recibido en la URL
    marca = get_object_or_404(Marca, nombre=marca_nombre)
    # Filtrar los materiales por la marca encontrada
    materiales = Material.objects.filter(marca=marca)

    # Configuración de la paginación
    paginator = Paginator(materiales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_materiales_filtrados_dep.html'
    else:
        template_name = 'ABM/materiales/ver_materiales_filtrados.html'

    return render(request, template_name,{'materiales': materiales, 'page_obj': page_obj, 'filtro': 'marca'})


def ver_materiales_proveedores(request, proveedor):
    id_proveedor = Proveedor.objects.get(nombre=proveedor)
    materiales = Material.objects.filter(id_proveedor=id_proveedor)
    # Configuración de la paginación
    paginator = Paginator(materiales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_materiales_filtrados_dep.html'
    else:
        template_name = 'ABM/materiales/ver_materiales_filtrados.html'

    return render(request, template_name, {'materiales': materiales, 'page_obj': page_obj, 'filtro': 'proveedor'})


def ver_materiales_stock(request, cantidad):
    if cantidad == 'Menos de 10':
        materiales = Material.objects.filter(unidades_stock__lt=10)
    elif cantidad == 'Menos de 50':
        materiales = Material.objects.filter(unidades_stock__lt=50)
    elif cantidad == 'Menos de 100':
        materiales = Material.objects.filter(unidades_stock__lt=100)
    # Configuración de la paginación
    paginator = Paginator(materiales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_materiales_filtrados_dep.html'
    else:
        template_name = 'ABM/materiales/ver_materiales_filtrados.html'

    return render(request, template_name, {'materiales': materiales, 'page_obj': page_obj, 'filtro': 'unidades en stock'})

def ver_materiales_faltantes(request):
    # Paso 1: Obtener los materiales requeridos para los pedidos pendientes
    materiales_requeridos = defaultdict(int)
    pedidos_pendientes = Pedido.objects.filter(estado='P')
    for pedido in pedidos_pendientes:
        for material_pedido in pedido.materialpedido_set.all():
            cantidad_requerida = material_pedido.cantidad + material_pedido.material.minimo
            materiales_requeridos[material_pedido.material] += cantidad_requerida

    # Paso 2: Verificar si hay suficiente stock para cada material y si ya hay un pedido pendiente
    materiales_a_comprar = []
    for material, cantidad_requerida in materiales_requeridos.items():
        stock_disponible = material.unidades_stock - material.minimo
        if stock_disponible < cantidad_requerida:
            # Verificar si ya existe un pedido pendiente para el material faltante
            pedido_compra_existente = PedidoCompra.objects.filter(materialpedidocompra__material=material, estado='P').first()
            if not pedido_compra_existente:
                # Si no hay suficiente stock y no hay un pedido pendiente, agregar el material a la lista de compra
                materiales_a_comprar.append((material, cantidad_requerida))

    # Paso 3: Crear pedidos de compra para materiales faltantes
    for material, cantidad_comprar in materiales_a_comprar:
        # Crear un pedido de compra para el material faltante
        pedido_compra = PedidoCompra.objects.create(fecha_solicitud=date.today(), estado='P')
        # Registrar el material y la cantidad en el pedido de compra
        MaterialPedidoCompra.objects.create(pedido_compra=pedido_compra, material=material, cantidad=cantidad_comprar)
        # Actualizar el estado del material a "falta_stock"
        material.falta_stock = True
        material.save()

    # Obtener los materiales que están en los pedidos y no están en stock
    materiales_faltantes = Material.objects.filter(pk__in=[material.pk for material, _ in materiales_a_comprar])

    # Obtener los pedidos de compra pendientes
    pedidos_compra_pendientes = PedidoCompra.objects.filter(estado='P')
    print(materiales_faltantes)
    print(pedidos_compra_pendientes)
    return render(request, 'pantallas_deposito/ver_materiales_faltantes.html', {
        'materiales_faltantes': materiales_faltantes,
        'pedidos_compra_pendientes': pedidos_compra_pendientes,
    })

def agregar_pedido_compra(request):
    if request.method == 'POST':
        nombre_producto = request.POST.get('nombre_producto')
        marca_id = request.POST.get('marca')
        otra_marca = request.POST.get('otra_marca')
        unidad_medida_id = request.POST.get('unidad_medida')
        otra_unidad_medida = request.POST.get('otra_unidad_medida')
        cantidad = request.POST.get('cantidad')

        pedido_compra = PedidoCompra.objects.create(
            fecha_solicitud=date.today(),
            fecha_entrega=None,
            estado='P'
        )

        if marca_id == "otra":
            marca_nueva = Marca.objects.create(nombre=otra_marca)
            marca = marca_nueva
        else:
            marca = Marca.objects.get(id=marca_id)

        if unidad_medida_id == "otra":
            unidad_medida_nueva = UnidadMedida.objects.create(descripcion=otra_unidad_medida)
            unidad_medida = unidad_medida_nueva
        else:
            unidad_medida = UnidadMedida.objects.get(id=unidad_medida_id)

        material = Material.objects.create(
            nombre=nombre_producto,
            marca=marca,
            medida=unidad_medida
        )

        material_pedido_compra = MaterialPedidoCompra.objects.create(
            pedido_compra=pedido_compra,
            material=material,
            cantidad=cantidad
        )
        print(f'se creo el pedido compra {material_pedido_compra}')

        return redirect('pantallas_deposito/ver_materiales_faltantes')

    return render(request, 'pantallas_deposito/ver_materiales_faltantes.html')


def exportar_excel(request):
    # Crear un nuevo libro de Excel y una nueva hoja
    wb = Workbook()
    ws = wb.active
    tipo_dato = request.POST.get('tipo_dato')  # Por ejemplo, 'Ingenieros', 'Proyectos', etc.
    criterio = request.POST.get('criterio')  # Por ejemplo, el nombre del cliente

    if tipo_dato == 'Ingenieros':
        ws.title = "Ingenieros"
        # Añadir encabezados a la hoja
        encabezados = ['ID', 'Username', 'First Name', 'Last Name', 'Email']
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = encabezado
            ws.column_dimensions[col_letter].width = 15

        # Obtener datos de los ingenieros
        group_ingeniero = Group.objects.get(name='INGENIERO')
        ingenieros = group_ingeniero.user_set.all()
        # Añadir datos a la hoja
        for ingeniero in ingenieros:
            ws.append([ingeniero.id, ingeniero.username, ingeniero.first_name, ingeniero.last_name, ingeniero.email])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Ingenieros.xlsx'
        wb.save(response)

    elif tipo_dato == 'Clientes':
        ws.title = "Clientes"
        # Añadir encabezados a la hoja
        encabezados = ['ID', 'Tipo_persona', 'Nombre', 'RUC', 'Telefono', 'Email', 'Ciudad', 'Direccion',
                       'Observaciones']
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = encabezado
            ws.column_dimensions[col_letter].width = 15

        # Obtener datos de los clientes
        clientes = Cliente.objects.all()

        # Añadir datos a la hoja
        for cliente in clientes:
            ws.append([cliente.id, cliente.tipo_persona, cliente.nombre, cliente.ruc, cliente.telefono, cliente.email,
                       cliente.ciudad.nombre, cliente.direccion, cliente.observaciones])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Clientes.xlsx'
        wb.save(response)

    elif tipo_dato == 'Proveedores':

        ws.title = "Proveedores"
        encabezados = ['ID', 'Nombre', 'RUC', 'Telefono', 'Email', 'Pagina web', 'Ciudad', 'Direccion', 'Observaciones']
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = encabezado
            ws.column_dimensions[col_letter].width = 15
        proveedores = Proveedor.objects.all()
        for proveedor in proveedores:
            ws.append([proveedor.id, proveedor.nombre, proveedor.ruc, proveedor.telefono, proveedor.email,
                       proveedor.pagina_web, proveedor.ciudad.nombre, proveedor.direccion, proveedor.observaciones])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Proveedores.xlsx'
        wb.save(response)
    elif tipo_dato == 'Materiales':
        ws.title = "Materiales"
        # Añadir encabezados a la hoja
        encabezados = ['ID', 'Nombre', 'Marca', 'Proveedor', 'Medida', 'Mínimo', 'Unidades en Stock']
        for col_num, encabezado in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = encabezado
            ws.column_dimensions[col_letter].width = 15

        # Obtener datos de los materiales
        materiales = Material.objects.all().select_related(
            'id_proveedor')  # select_related para optimizar la consulta relacionada con Proveedor

        # Añadir datos a la hoja
        for material in materiales:
            ws.append([material.id, material.nombre, material.marca, material.id_proveedor.nombre, material.medida,
                       material.minimo, material.unidades_stock])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Materiales.xlsx'
        wb.save(response)

    return response


def exportar_pdf(request):
    tipo_dato = request.POST.get('tipo_dato')  # Por ejemplo, 'Ingenieros', 'Proyectos', etc.
    criterio = request.POST.get('criterio')  # Por ejemplo, el nombre del cliente
    # Crear un nuevo documento PDF
    response = HttpResponse(content_type='application/pdf')

    if tipo_dato == 'Ingenieros':
        response['Content-Disposition'] = 'attachment; filename="Ingenieros.pdf"'
        doc = SimpleDocTemplate(response, pagesize=A4)
        # Obtener datos de los ingenieros
        group_ingeniero = Group.objects.get(name='INGENIERO')
        ingenieros = group_ingeniero.user_set.all()
        data = [['ID', 'Usuario', 'Nombre', 'Apellido', 'Email']]
        for ingeniero in ingenieros:
            data.append([ingeniero.id, ingeniero.username, ingeniero.first_name, ingeniero.last_name, ingeniero.email])

        # Crear una tabla con los datos
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        # Construir el documento
        story = [table]
        doc.build(story)

    elif tipo_dato == 'Clientes':
        response['Content-Disposition'] = 'attachment; filename="Clientes.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        # Datos para la tabla
        encabezados = ['ID', 'Tipo_persona', 'Nombre', 'RUC', 'Telefono', 'Email', 'Ciudad', 'Direccion',
                       'Observaciones']

        # Obtener datos de los clientes
        clientes = Cliente.objects.all()
        for cliente in clientes:
            data.append([cliente.id, cliente.tipo_persona, cliente.nombre, cliente.ruc, cliente.telefono, cliente.email,
                         cliente.ciudad.nombre,
                         cliente.direccion, cliente.observaciones])

        # Crear tabla
        table = Table(data)

        # Aplicar estilos a la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)

        # Alternar colores de fondo para las filas
        rowNumb = len(data)
        for i in range(1, rowNumb):
            if i % 2 == 0:
                bc = colors.bisque
            else:
                bc = colors.beige
            ts = TableStyle(
                [('BACKGROUND', (0, i), (-1, i), bc)]
            )
            table.setStyle(ts)

        # Agregar tabla al documento
        elems = []
        elems.append(table)

        doc.build(elems)
    elif tipo_dato == 'Proveedores':
        response['Content-Disposition'] = 'attachment; filename="Proveedores.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        # Datos para la tabla
        data = [['ID', 'Nombre', 'RUC', 'Telefono', 'Email', 'Ciudad', 'Direccion', 'Página Web', 'Observaciones']]

        # Obtener datos de los proveedores
        proveedores = Proveedor.objects.all()
        for proveedor in proveedores:
            data.append(
                [proveedor.id, proveedor.nombre, proveedor.ruc, proveedor.telefono, proveedor.email,
                 proveedor.ciudad.nombre, proveedor.direccion,
                 proveedor.pagina_web, proveedor.observaciones])

        # Crear tabla
        table = Table(data)

        # Aplicar estilos a la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)

        # Alternar colores de fondo para las filas
        rowNumb = len(data)
        for i in range(1, rowNumb):
            if i % 2 == 0:
                bc = colors.bisque
            else:
                bc = colors.beige
            ts = TableStyle(
                [('BACKGROUND', (0, i), (-1, i), bc)]
            )
            table.setStyle(ts)

        # Agregar tabla al documento
        elems = []
        elems.append(table)

        doc.build(elems)
    elif tipo_dato == 'Materiales':
        response['Content-Disposition'] = 'attachment; filename="Materiales.pdf"'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        # Datos para la tabla
        data = [['ID', 'Nombre', 'Marca', 'Proveedor', 'Medida', 'Mínimo', 'Unidades en Stock']]

        # Obtener datos de los materiales
        materiales = Material.objects.all().select_related(
            'id_proveedor')  # select_related para optimizar la consulta relacionada con Proveedor
        for material in materiales:
            data.append(
                [material.id, material.nombre, material.marca, material.id_proveedor.nombre, material.medida,
                 material.minimo, material.unidades_stock])

        # Crear tabla
        table = Table(data)

        # Aplicar estilos a la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)

        # Alternar colores de fondo para las filas
        rowNumb = len(data)
        for i in range(1, rowNumb):
            if i % 2 == 0:
                bc = colors.bisque
            else:
                bc = colors.beige
            ts = TableStyle(
                [('BACKGROUND', (0, i), (-1, i), bc)]
            )
            table.setStyle(ts)

        # Agregar tabla al documento
        elems = []
        elems.append(table)

        doc.build(elems)
    return response


def extraer_plazo(sheet):
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell and "Plazo de ejecución" in str(cell):
                match = re.search(r'Plazo de ejecución: (\d+) días', cell)
                if match:
                    return int(match.group(1))
    return None


def cargar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    if request.method == "POST" and request.FILES['archivo_presupuesto']:
        archivo_excel = request.FILES['archivo_presupuesto']
        # No necesitas la ruta completa aquí, simplemente el nombre del archivo.
        # Django se encargará de guardar el archivo en la ubicación correcta según tu configuración de MEDIA_ROOT.
        nombre_archivo = os.path.basename(archivo_excel.name)

        # Crear una instancia de ArchivoPresupuesto
        archivo_presupuesto = ArchivoPresupuesto(presupuesto=presupuesto, nombre=nombre_archivo)

        # Guardar el archivo en el modelo
        archivo_presupuesto.archivo.save(nombre_archivo, archivo_excel)

        # Guardar el modelo en la base de datos

        workbook = load_workbook(archivo_presupuesto.archivo)
        sheet = workbook.active

        # extraemos el plazo de ejecucion
        plazo = extraer_plazo(sheet)
        if plazo:
            obra = presupuesto.proyecto.obra
            obra.plazo = plazo
            obra.save()
            print(plazo)

        # Variables para identificar el inicio de la tabla
        inicio_tabla_fila = None
        inicio_tabla_columna = None

        # Navegar por cada fila
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            found_rubros = False
            for cell in row:
                if cell and "Rubros" == cell.strip():
                    found_rubros = True
                    break
            if found_rubros:
                # Obtener el índice de la celda que contiene "Rubros"
                rubros_index = row.index(cell)
                # Si la siguiente celda es "Un", entonces hemos encontrado el inicio de la tabla
                if row[rubros_index + 1].strip() == "Un":
                    inicio_tabla_fila = idx
                    inicio_tabla_columna = rubros_index
                    print(
                        f'Inicio de tabla identificado en la fila {inicio_tabla_fila} y columna {inicio_tabla_columna}')
                    break

        # Si encontramos el inicio de la tabla, procesamos los datos
        if inicio_tabla_fila is not None and inicio_tabla_columna is not None:
            categoria_actual = None
            item_actual = None

            idx = inicio_tabla_fila + 2  # Comenzamos en la fila siguiente a inicio_tabla

            while idx < sheet.max_row and sheet[idx][
                inicio_tabla_columna].value:  # Mientras no lleguemos al final del archivo
                fila_actual = [cell.value for cell in sheet[idx]]
                valor_rubros = fila_actual[inicio_tabla_columna].strip() if fila_actual[inicio_tabla_columna] else None
                valor_un = fila_actual[inicio_tabla_columna + 1].strip() if fila_actual[
                    inicio_tabla_columna + 1] else None
                # Comprobar si la fila siguiente tiene valor en "Un"
                fila_siguiente = [cell.value for cell in sheet[idx + 1]] if idx + 1 <= sheet.max_row else None
                valor_un_siguiente = fila_siguiente[inicio_tabla_columna + 1].strip() if fila_siguiente and \
                                                                                         fila_siguiente[
                                                                                             inicio_tabla_columna + 1] else None

                if valor_rubros and not valor_un:
                    if valor_un_siguiente:
                        # Es un Item
                        item_actual = Item(nombre=valor_rubros, categoria=categoria_actual, archivo=archivo_presupuesto)
                        item_actual.save()
                        # NO limpiamos la categoría actual aquí
                    else:
                        # Es una Categoría
                        categoria_actual = Categoria(nombre=valor_rubros, archivo=archivo_presupuesto)
                        categoria_actual.save()
                        item_actual = None  # Limpiamos el ítem actual
                elif valor_rubros and valor_un:
                    # Es un SubItem
                    subitem = SubItem(
                        item=item_actual,
                        rubro=valor_rubros,
                        unidad_medida=valor_un,
                        cantidad=fila_actual[inicio_tabla_columna + 2],
                        precio_unitario=fila_actual[inicio_tabla_columna + 3],
                        precio_total=(fila_actual[inicio_tabla_columna + 2] * fila_actual[inicio_tabla_columna + 3])
                    )
                    subitem.save()
                else:
                    # Aquí puedes manejar casos no previstos o simplemente ignorarlos
                    pass

                idx += 1  # Avanzar a la siguiente fila

        else:
            print("No se encontró la tabla en el archivo proporcionado.")

        workbook.close()
        # Calcular la suma de todos los campos `precio_total` de los subitems que pertenecen al presupuesto en cuestión, y cargarlas como monto_total del presupuesto
        suma_total = \
        SubItem.objects.filter(item__categoria__archivo__presupuesto=presupuesto).aggregate(suma=Sum('precio_total'))[
            'suma']
        presupuesto.monto_total = suma_total
        presupuesto.monto_anticipo = suma_total / 2
        presupuesto.save()

        # Redirige al usuario a donde desees luego de procesar el archivo
        return redirect('ver_presupuestos')
    # Redirige al usuario a donde desees luego de procesar el archivo
    return redirect('ver_presupuestos')


def ver_archivo_presupuesto(request, pk):
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=pk)
    presupuesto = Presupuesto.objects.get(id=pk)
    # Estructura para mantener la jerarquía
    estructura_presupuesto = []
    subtotal = presupuesto.subtotal
    iva = presupuesto.iva
    monto_total = presupuesto.monto_total
    cliente = presupuesto.proyecto.cliente.nombre
    proyecto = presupuesto.proyecto.nombre
    # Iterar sobre las secciones asociadas con el archivo de presupuesto
    for seccion in archivo_presupuesto.secciones.all():
        subsecciones_data = []
        # Buscar subsecciones asociadas a esta sección y que estén en ArchivoPresupuesto
        subsecciones = seccion.subseccion_set.filter(
            id__in=archivo_presupuesto.subsecciones.values_list('id', flat=True))
        for subseccion in subsecciones:
            # Filtrar detalles asociados a esta subsección y que estén en ArchivoPresupuesto
            detalles = subseccion.detalle_set.filter(id__in=archivo_presupuesto.detalles.values_list('id', flat=True))
            subsecciones_data.append((subseccion, detalles))

        estructura_presupuesto.append((seccion, subsecciones_data))

    context = {
        'estructura_presupuesto': estructura_presupuesto,
        'subtotal': subtotal,
        'monto_total': monto_total,
        'iva': iva,
        'proyecto': proyecto,
        'cliente': cliente,
        'presupuesto_id': presupuesto.id,
        'archivo_presupuesto_id': archivo_presupuesto.id
    }

    return render(request, 'pantallas_ing/ver_presupuesto.html', context)


def modificar_presupuesto(request, pk):
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=pk)
    categorias = Categoria.objects.filter(archivo=archivo_presupuesto).order_by('pk')
    presupuesto = archivo_presupuesto.presupuesto
    # Pasamos el archivo_presupuesto y las categorías al template
    context = {
        'archivo_presupuesto': archivo_presupuesto,
        'categorias': categorias,
        'presupuesto': presupuesto
    }
    return render(request, 'pantallas_ing/modificar_presupuesto.html', context)


def editar_categoria(request, categoria_id):
    # Lógica para editar la categoría
    return JsonResponse({'status': 'success'})


def editar_item(request, item_id):
    # Lógica para editar el item
    return JsonResponse({'status': 'success'})


# Función para actualizar el monto_total de un presupuesto basado en sus subitems
def actualizar_monto_total(presupuesto):
    suma_total = \
    SubItem.objects.filter(item__categoria__archivo__presupuesto=presupuesto).aggregate(suma=Sum('precio_total'))[
        'suma']
    presupuesto.monto_total = suma_total
    presupuesto.monto_anticipo = suma_total / 2
    presupuesto.save()


@csrf_exempt
def editar_subitem(request, subitem_id):
    if request.method == "POST":
        data = json.loads(request.body)
        try:
            subitem = SubItem.objects.get(pk=subitem_id)
            subitem.rubro = data.get('rubro')
            subitem.unidad_medida = data.get('unidad_medida')
            subitem.cantidad = Decimal(data.get('cantidad'))
            subitem.precio_unitario = Decimal(data.get('precio_unitario'))
            subitem.precio_total = subitem.cantidad * subitem.precio_unitario
            subitem.save()

            presupuesto_asociado = subitem.item.categoria.archivo.presupuesto
            actualizar_monto_total(presupuesto_asociado)

            return JsonResponse({'status': 'success'})

        except SubItem.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Subitem no encontrado'}, status=404)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@csrf_exempt
def eliminar_subitem(request, subitem_id):
    if request.method == "DELETE":
        try:
            subitem = SubItem.objects.get(pk=subitem_id)
            presupuesto_asociado = subitem.item.categoria.archivo.presupuesto
            # Capturamos el Item asociado al SubItem antes de eliminarlo
            item_asociado = subitem.item
            subitem.delete()

            # Verificar si el Item no tiene otros SubItems asociados
            if not item_asociado.subitem_set.exists():
                # Capturamos la Categoría asociada al Item antes de eliminarlo
                categoria_asociada = item_asociado.categoria
                item_asociado.delete()

                # Verificar si la Categoría no tiene otros Items asociados
                if not categoria_asociada.item_set.exists():
                    categoria_asociada.delete()

            actualizar_monto_total(presupuesto_asociado)
            return JsonResponse({'status': 'success'})

        except SubItem.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Subitem no encontrado'}, status=404)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


@csrf_exempt
def actualizar_estado(request, presupuesto_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            presupuesto = Presupuesto.objects.get(pk=presupuesto_id)
            presupuesto.estado = nuevo_estado
            presupuesto.save()
            return JsonResponse({'status': 'success'})
        except Presupuesto.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Presupuesto no encontrado'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def ver_presupuestos_adm(request):
    # Si hay filtros en la petición, úsalos.
    filtro_campo = request.GET.get('campo', None)
    filtro_valor = request.GET.get('valor', None)

    presupuestos = Presupuesto.objects.all()

    if filtro_campo and filtro_valor:
        if filtro_campo == 'cliente':
            presupuestos = presupuestos.filter(proyecto__cliente__nombre=filtro_valor)
        elif filtro_campo == 'encargadoPresupuesto':
            ingeniero = User.objects.get(username=filtro_valor)
            presupuestos = presupuestos.filter(encargado=ingeniero)
        elif filtro_campo == 'estadoPresupuesto':
            estado_codigo = {value: key for key, value in dict(Presupuesto.ESTADOS).items()}[filtro_valor]
            presupuestos = presupuestos.filter(estado=estado_codigo)
        elif filtro_campo == 'estadoAnticipo':
            tiene_anticipo = True if filtro_valor == 'Sí' else False
            presupuestos = presupuestos.filter(anticipo=tiene_anticipo)

    # Lógica de paginación
    paginator = Paginator(presupuestos, 10)  # 10 presupuestos por página
    page = request.GET.get('page')

    try:
        presupuestos_paginados = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, muestra la primera página.
        presupuestos_paginados = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera de rango, muestra la última página de resultados.
        presupuestos_paginados = paginator.page(paginator.num_pages)

    context = {
        'page_obj': presupuestos_paginados,  # Usa 'page_obj' en lugar de 'presupuestos' en el template
        'campo_seleccionado': filtro_campo,
        'valor_seleccionado': filtro_valor,
        'presupuestos': presupuestos
    }

    return render(request, 'pantallas_adm/ver_presupuestos_adm.html', context)


@csrf_exempt
def actualizar_anticipo(request, presupuesto_id):
    if request.method == "POST":
        try:
            presupuesto = Presupuesto.objects.get(pk=presupuesto_id)
            logger.debug(f'Presupuesto encontrado: {presupuesto}')
            presupuesto.anticipo = request.POST.get('anticipo') == 'true'
            logger.debug(f'Anticipo valor: {presupuesto.anticipo}')

            # Si se ha enviado un archivo de comprobante, lo guardamos
            comprobante = request.FILES.get('comprobante')
            if comprobante:
                presupuesto.comprobante_anticipo = comprobante
                logger.debug('Comprobante guardado')

            # Verificar si monto_total tiene un valor válido
            if presupuesto.monto_total is not None:
                # Calcular el monto_anticipo como el 50% del monto_total
                fifty_percent = Decimal('0.5')
                presupuesto.monto_anticipo = (presupuesto.monto_total * fifty_percent).quantize(presupuesto.monto_total)
                logger.debug(f'Monto anticipo calculado: {presupuesto.monto_anticipo}')

                if presupuesto.estado == 'S' and presupuesto.anticipo is True:
                    presupuesto.estado = 'A'
                if presupuesto.anticipo is True:
                    presupuesto.fecha_pago_anticipo = date.today()
                presupuesto.save()
                logger.debug('Presupuesto guardado')

                return JsonResponse({
                    'status': 'success',
                    'nuevo_estado': presupuesto.get_estado_display(),
                    'fecha_pago_anticipo': presupuesto.fecha_pago_anticipo.strftime('%d/%m/%Y') if presupuesto.fecha_pago_anticipo else None
                })
            else:
                logger.error('El monto total no está definido')
                return JsonResponse({'status': 'error', 'message': 'El monto total no está definido'})
        except Exception as e:
            logger.error(f'Error al actualizar el anticipo: {str(e)}')
            return JsonResponse({'status': 'error', 'message': str(e)})

    logger.error('Método no permitido')
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def ver_obras_adm(request):
    obras = Obra.objects.filter(encargado__isnull=False)
    presupuestos_con_anticipo = Presupuesto.objects.filter(anticipo=True)
    # Filtramos aquellos presupuestos cuyo proyecto no tiene una obra con encargado asignado
    presupuestos_a_asignar = [presupuesto for presupuesto in presupuestos_con_anticipo if
                              not presupuesto.proyecto.obra.encargado]

    # Obtenemos los ingenieros disponibles
    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.all()

    # Obtenemos los meses de las fechas de inicio y fin
    meses_inicio = list(set(obras.values_list('fecha_inicio__month', flat=True)))
    meses_fin = list(set(obras.values_list('fecha_fin__month', flat=True)))
    paginator = Paginator(obras, 10)  # Muestra 10 obras por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'obras': obras,
        'presupuestos': presupuestos_a_asignar,
        'ingenieros': ingenieros,
        'meses_inicio': meses_inicio,
        'meses_fin': meses_fin,
        'page_obj': page_obj
    }
    template_name = 'pantallas_adm/ver_obras_adm.html'

    return render(request, template_name, context)





def ver_obra_adm(request, obra_id):
    obra = get_object_or_404(Obra, id=obra_id)
    return render(request, 'pantallas_adm/ver_obra_adm.html', {'obra': obra})


def buscar_obras(request):
    q = request.GET.get('q', '')
    obras = Obra.objects.filter(proyecto__nombre__icontains=q).values('id', 'proyecto__nombre')
    return JsonResponse(list(obras), safe=False)


def obtener_estados_obra(request):
    # Obtenemos los estados distintos del modelo Obra
    estados_distintos = Obra.objects.values_list('estado', flat=True).distinct()

    # Creamos una instancia ficticia de Obra para poder usar el método get_estado_display()
    obra_dummy = Obra()

    # Obtenemos la representación legible para cada estado
    estados_legibles = [obra_dummy.get_estado_display() for obra_dummy.estado in estados_distintos]

    return JsonResponse(estados_legibles, safe=False)


def obtener_ingenieros_obra(request):
    # Obtener ingenieros de las obras
    ingenieros = Obra.objects.exclude(encargado=None).values_list('encargado__first_name',
                                                                  'encargado__last_name').distinct()

    # Convertir la tupla (first_name, last_name) a la cadena "first_name last_name"
    formatted_ingenieros = [f"{ingeniero[0]} {ingeniero[1]}" for ingeniero in ingenieros]

    return JsonResponse(formatted_ingenieros, safe=False)


def obtener_fechas_inicio(request):
    fechas = Obra.objects.exclude(fecha_inicio=None).dates('fecha_inicio', 'month').distinct()
    fechas_formateadas = ["{} {}".format(traducir_mes(fecha.strftime('%B')), fecha.year) for fecha in fechas]
    return JsonResponse(fechas_formateadas, safe=False)


def obtener_fechas_fin(request):
    fechas = Obra.objects.exclude(fecha_fin=None).dates('fecha_fin', 'month').distinct()
    fechas_formateadas = ["{} {}".format(traducir_mes(fecha.strftime('%B')), fecha.year) for fecha in fechas]
    return JsonResponse(fechas_formateadas, safe=False)


def traducir_mes(mes_ingles):
    """Traduce un mes de inglés a español."""
    traducciones = {
        'January': 'Enero',
        'February': 'Febrero',
        'March': 'Marzo',
        'April': 'Abril',
        'May': 'Mayo',
        'June': 'Junio',
        'July': 'Julio',
        'August': 'Agosto',
        'September': 'Septiembre',
        'October': 'Octubre',
        'November': 'Noviembre',
        'December': 'Diciembre',
    }
    return traducciones.get(mes_ingles, mes_ingles)


def traducir_mes_inverso(mes_espanol):
    """Traduce un mes de español a inglés."""
    traducciones = {
        'Enero': 'January',
        'Febrero': 'February',
        'Marzo': 'March',
        'Abril': 'April',
        'Mayo': 'May',
        'Junio': 'June',
        'Julio': 'July',
        'Agosto': 'August',
        'Septiembre': 'September',
        'Octubre': 'October',
        'Noviembre': 'November',
        'Diciembre': 'December',
    }
    return traducciones.get(mes_espanol, mes_espanol)


def estado_legible_a_codigo(estado_legible):
    # Usamos el diccionario invertido para obtener el código a partir de la representación legible
    estado_codigo = dict((v, k) for k, v in Obra.ESTADOS).get(estado_legible)
    return estado_codigo


def ver_obras_filtradas(request):
    campo = request.GET.get('campo')
    valor = request.GET.get('valor')

    if campo == 'estado':
        # Convertimos el estado legible a su código correspondiente
        estado_codigo = estado_legible_a_codigo(valor)
        obras = Obra.objects.filter(estado=estado_codigo, encargado__isnull=False)
    elif campo == 'encargado':
        # Divide el valor en nombre y apellido
        first_name, last_name = valor.split(' ', 1)

        # Obtiene el usuario por nombre y apellido
        ingeniero = User.objects.get(first_name=first_name, last_name=last_name)

        # Filtra las obras por el ID del ingeniero
        obras = Obra.objects.filter(encargado=ingeniero)
    elif campo == 'fechaInicio':
        if valor == 'No iniciada':
            obras = Obra.objects.filter(fecha_inicio__isnull=True, encargado__isnull=False)
        else:
            # Convertir el mes en formato textual a un valor numérico
            month_es, year = valor.split()
            month_en = traducir_mes_inverso(month_es)
            month_num = datetime.strptime(month_en, '%B').month
            obras = Obra.objects.filter(fecha_inicio__month=month_num, fecha_inicio__year=year, encargado__isnull=False)
    elif campo == 'fechaFin':
        if valor == 'No finalizada':
            obras = Obra.objects.filter(fecha_fin__isnull=True, encargado__isnull=False)
        else:
            # Convertir el mes en formato textual a un valor numérico
            month_es, year = valor.split()
            month_en = traducir_mes_inverso(month_es)
            month_num = datetime.datetime.strptime(month_en, '%B').month
            obras = Obra.objects.filter(fecha_fin__month=month_num, fecha_fin__year=year, encargado__isnull=False)
    else:
        obras = Obra.objects.filter(encargado__isnull=False)

    presupuestos_con_anticipo = Presupuesto.objects.filter(anticipo=True)
    presupuestos_a_asignar = [presupuesto for presupuesto in presupuestos_con_anticipo if
                              not presupuesto.proyecto.obra.encargado]

    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.all()

    meses_inicio = list(set(obras.values_list('fecha_inicio__month', flat=True)))
    meses_fin = list(set(obras.values_list('fecha_fin__month', flat=True)))
    paginator = Paginator(obras, 10)  # Muestra 10 obras por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'obras': obras,
        'presupuestos': presupuestos_a_asignar,
        'ingenieros': ingenieros,
        'meses_inicio': meses_inicio,
        'meses_fin': meses_fin,
        'campo_seleccionado': campo,
        'valor_seleccionado': valor,
        'page_obj': page_obj
    }

    return render(request, 'pantallas_adm/ver_obras_adm.html', context)


def asignar_obras(request):
    # Buscamos los presupuestos que tienen anticipo pagado
    presupuestos_con_anticipo = Presupuesto.objects.filter(anticipo=True)
    # Filtramos aquellos presupuestos cuyo proyecto no tiene una obra con encargado asignado
    presupuestos_a_asignar = [presupuesto for presupuesto in presupuestos_con_anticipo if
                              not presupuesto.proyecto.obra.encargado]
    group_ingeniero = Group.objects.get(name='INGENIERO')
    ingenieros = group_ingeniero.user_set.all()
    context = {'presupuestos': presupuestos_a_asignar,
               'ingenieros': ingenieros}
    return render(request, 'pantallas_adm/asignar_obras.html', context)


@csrf_exempt
def asignar_ingeniero_a_obra(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            presupuesto_id = data.get('presupuesto_id')
            ingeniero_id = data.get('ingeniero_id')
            print(presupuesto_id)
            # Obtener el presupuesto
            presupuesto = Presupuesto.objects.get(pk=presupuesto_id)
            print(presupuesto)
            # Obtener la obra relacionada con el proyecto del presupuesto
            obra = Obra.objects.get(proyecto=presupuesto.proyecto)

            # Asignar el ingeniero a la obra
            obra.encargado = User.objects.get(pk=ingeniero_id)
            obra.save()

            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def buscar_presupuestos(request):
    query = request.GET.get('q', '')
    presupuestos = Presupuesto.objects.filter(
        Q(proyecto__nombre__icontains=query)
    )
    data = [
        {
            'id': presupuesto.id,
            'proyecto_nombre': presupuesto.proyecto.nombre,
            'cliente_nombre': presupuesto.proyecto.cliente.nombre
        }
        for presupuesto in presupuestos
    ]
    return JsonResponse(data, safe=False)


def buscar_presupuestos_terminados(request):
    query = request.GET.get('q', '')
    presupuestos = Presupuesto.objects.filter(
        Q(proyecto__nombre__icontains=query),
        Q(estado__in=['S', 'A']),
        Q(anticipo=False)
    )
    data = [
        {
            'id': presupuesto.id,
            'proyecto_nombre': presupuesto.proyecto.nombre,
            'cliente_nombre': presupuesto.proyecto.cliente.nombre,
            'monto_total': str(presupuesto.monto_total)  # Convertir a string para serialización JSON
        }
        for presupuesto in presupuestos
    ]
    return JsonResponse(data, safe=False)


def ver_presupuesto_adm(request, presupuesto_id):
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=presupuesto_id)
    presupuesto = Presupuesto.objects.get(id=presupuesto_id)
    # Estructura para mantener la jerarquía
    estructura_presupuesto = []
    subtotal = presupuesto.subtotal
    iva = presupuesto.iva
    monto_total = presupuesto.monto_total
    cliente = presupuesto.proyecto.cliente.nombre
    proyecto = presupuesto.proyecto.nombre
    # Iterar sobre las secciones asociadas con el archivo de presupuesto
    for seccion in archivo_presupuesto.secciones.all():
        subsecciones_data = []
        # Buscar subsecciones asociadas a esta sección y que estén en ArchivoPresupuesto
        subsecciones = seccion.subseccion_set.filter(
            id__in=archivo_presupuesto.subsecciones.values_list('id', flat=True))
        for subseccion in subsecciones:
            # Filtrar detalles asociados a esta subsección y que estén en ArchivoPresupuesto
            detalles = subseccion.detalle_set.filter(id__in=archivo_presupuesto.detalles.values_list('id', flat=True))
            subsecciones_data.append((subseccion, detalles))

        estructura_presupuesto.append((seccion, subsecciones_data))

    context = {
        'estructura_presupuesto': estructura_presupuesto,
        'subtotal': subtotal,
        'monto_total': monto_total,
        'iva': iva,
        'proyecto': proyecto,
        'cliente': cliente,
        'presupuesto_id': presupuesto.id,
        'archivo_presupuesto_id': archivo_presupuesto.id,
        'presupuesto': presupuesto
    }
    return render(request, 'pantallas_adm/ver_presupuesto_adm.html', context)


def obtener_estados_anticipo(request):
    estados = ['Sí', 'No']
    return JsonResponse(list(estados), safe=False)


def ver_proyectos_estado_anticipo(request, anticipo):
    if anticipo == 'Sí':
        presupuestos = Presupuesto.objects.filter(anticipo=True)
    else:
        presupuestos = Presupuesto.objects.filter(anticipo=False)
    return render(request, 'pantallas_adm/ver_presupuestos_filtrados.html', {'presupuestos': presupuestos})


def ver_presupuestos_cliente(request, cliente_nombre):
    presupuestos = Presupuesto.objects.filter(proyecto__cliente__nombre=cliente_nombre)
    return render(request, 'pantallas_adm/ver_presupuestos_filtrados.html', {'presupuestos': presupuestos})


def ver_presupuestos_encargado_presupuesto(request, ingeniero_user):
    ingeniero = User.objects.filter(username=ingeniero_user)
    presupuestos = Presupuesto.objects.filter(encargado__in=ingeniero)
    return render(request, 'pantallas_adm/ver_presupuestos_filtrados.html', {'presupuestos': presupuestos})


def ver_presupuestos_estado_presupuesto(request, estado):
    # Convertir el nombre legible del estado a su código correspondiente
    estado_codigo = {value: key for key, value in dict(Presupuesto.ESTADOS).items()}[estado]

    # Filtrar los proyectos basándose en el código del estado
    presupuestos = Presupuesto.objects.filter(estado=estado_codigo)

    return render(request, 'pantallas_adm/ver_presupuestos_filtrados.html', {'presupuestos': presupuestos})


def verificar_obras_agendadas():
    # Obtener todas las obras que deberían haber empezado pero aún están en estado "No Iniciada"
    obras_a_iniciar = Obra.objects.filter(estado='NI', fecha_inicio__lte=date.today())

    for obra in obras_a_iniciar:
        obra.estado = 'E'  # Cambiar el estado a "En ejecución"
        obra.save()


def iniciar_obra(request, obra_id):
    obra = get_object_or_404(Obra, id=obra_id)

    if obra.estado == 'NI':
        obra.fecha_inicio = date.today()
        obra.estado = 'E'  # 'E' es el estado para "En ejecución"
        obra.save()

        return JsonResponse({'status': 'success', 'message': 'Obra iniciada con éxito'})

    return JsonResponse({'status': 'error', 'message': 'No se pudo iniciar la obra'})


def agendar_inicio_obra(request, obra_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        fecha_inicio = data.get('fecha_inicio')
        try:
            obra = Obra.objects.get(id=obra_id)
            obra.fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            obra.save()
            print('ya')
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def finalizar_obra(request, obra_id):
    if request.method == 'POST':
        try:
            obra = Obra.objects.get(id=obra_id)
            obra.fecha_fin = date.today()
            obra.estado = 'F'  # F para Finalizada
            obra.save()
            return JsonResponse({'status': 'success', 'message': 'Obra finalizada con éxito',
                                 'fecha_fin': obra.fecha_fin.strftime('%Y-%m-%d')})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def obtener_presupuesto_detalle(request, proyecto_id):
    try:
        # Buscamos el presupuesto basado en el proyecto_id
        presupuesto = Presupuesto.objects.get(proyecto_id=proyecto_id)

        # Obtenemos el archivo_presupuesto relacionado
        archivo_presupuesto = ArchivoPresupuesto.objects.get(presupuesto=presupuesto)

        # Recuperamos las categorías y los subitems
        categorias = Categoria.objects.filter(archivo=archivo_presupuesto).order_by('pk')

        data = []
        for categoria in categorias:
            items_data = []
            for item in categoria.item_set.all():
                subitems_data = []
                for subitem in item.subitem_set.all():
                    subitems_data.append({
                        'id': subitem.id,
                        'rubro': subitem.rubro,
                        'unidad_medida': subitem.unidad_medida,
                        'cantidad': float(subitem.cantidad),
                        'precio_unitario': float(subitem.precio_unitario),
                        'precio_total': float(subitem.precio_total),
                        'itemId': subitem.item.id,
                        'categoriaId': subitem.item.categoria.id
                    })
                items_data.append({
                    'id': item.id,
                    'nombre': item.nombre,
                    'subitems': subitems_data,
                    'categoriaId': item.categoria.id
                })
            data.append({
                'id': categoria.id,
                'nombre': categoria.nombre,
                'items': items_data
            })

        return JsonResponse(data, safe=False)
    except (Presupuesto.DoesNotExist, ArchivoPresupuesto.DoesNotExist):
        return JsonResponse({'error': 'Presupuesto no encontrado'}, status=404)


def registrar_anticipo(request):
    presupuestos = Presupuesto.objects.filter(estado__in=['S', 'A']).filter(anticipo=False)
    return render(request, 'pantallas_adm/registrar_anticipo.html', {'presupuestos': presupuestos})


def obtener_filtro_valores(request):
    tipo = request.GET.get('tipo')
    presupuestos = Presupuesto.objects.filter(estado__in=['S', 'A']).filter(anticipo=False)
    if tipo == 'cliente':
        proyectos = [presupuesto.proyecto for presupuesto in presupuestos]
        clientes = Cliente.objects.filter(proyecto__in=proyectos).distinct()
        data = [{'id': cliente.id, 'nombre': cliente.nombre} for cliente in clientes]
    elif tipo == 'encargadoPresupuesto':
        ingenieros = User.objects.filter(presupuesto__in=presupuestos).distinct()
        data = [{'id': ingeniero.id, 'nombre': f"{ingeniero.first_name} {ingeniero.last_name}"} for ingeniero in
                ingenieros]
    elif tipo == 'estadoPresupuesto':
        data = [
            {'id': 'S', 'nombre': 'Enviado'},
            {'id': 'A', 'nombre': 'Aprobado'}
        ]
    else:
        data = []

    return JsonResponse(data, safe=False)


def obtener_presupuestos_filtrados(request):
    campo = request.GET.get('campo')
    valor = request.GET.get('valor')

    if campo == 'cliente':
        presupuestos = Presupuesto.objects.filter(proyecto__cliente_id=valor, estado__in=['S', 'A'], anticipo=False)
    elif campo == 'encargadoPresupuesto':
        presupuestos = Presupuesto.objects.filter(encargado_id=valor, estado__in=['S', 'A'], anticipo=False)
    elif campo == 'estadoPresupuesto':
        presupuestos = Presupuesto.objects.filter(estado=valor)
    else:
        presupuestos = []

    data = [
        {
            'id': presupuesto.id,
            'nombre': presupuesto.proyecto.nombre
        }
        for presupuesto in presupuestos
    ]

    return JsonResponse(data, safe=False)


def obtener_monto_presupuesto(request):
    presupuesto_id = request.GET.get('id')
    presupuesto = Presupuesto.objects.get(id=presupuesto_id)
    data = {
        'monto_total': str(presupuesto.monto_total)  # Convertir a string para que sea serializable
    }
    return JsonResponse(data)




def buscar_presupuestos_ingeniero(request):
    query = request.GET.get('q', '')
    # Filtra los presupuestos basándose en el nombre del proyecto y el ingeniero asignado
    presupuestos = Presupuesto.objects.filter(
        Q(proyecto__nombre__icontains=query) & Q(encargado=request.user)
    )
    data = [
        {
            'id': presupuesto.id,
            'proyecto_nombre': presupuesto.proyecto.nombre,
            'cliente_nombre': presupuesto.proyecto.cliente.nombre
        }
        for presupuesto in presupuestos
    ]
    return JsonResponse(data, safe=False)


def obtener_clientes_presupuestos_ing(request):
    clientes = Proyecto.objects.values_list('cliente__nombre', flat=True).distinct()
    return JsonResponse(list(clientes), safe=False)


def crear_presupuesto(request, presupuesto_id):
    presupuesto = Presupuesto.objects.get(id=presupuesto_id)
    secciones = Seccion.objects.all()
    subsecciones = SubSeccion.objects.all()
    detalles = Detalle.objects.all()

    secciones_json = json.loads(serialize('json', secciones))
    subsecciones_json = json.loads(serialize('json', subsecciones))
    detalles_json = json.loads(serialize('json', detalles))

    contexto = {
        'presupuesto': presupuesto,
        'secciones_json': json.dumps([s['fields']['nombre'] for s in secciones_json]),
        'subsecciones_json': json.dumps([ss['fields']['nombre'] for ss in subsecciones_json]),
        'detalles_json': json.dumps([d['fields']['rubro'] for d in detalles_json]),
    }

    return render(request, 'pantallas_ing/crear_presupuesto.html', contexto)


def obtener_subsecciones(request):
    # Sorted by 'nombre' field in ascending order
    subsecciones = SubSeccion.objects.all().order_by('nombre')
    data = [{'id': subseccion.id, 'nombre': subseccion.nombre} for subseccion in subsecciones]
    return JsonResponse(data, safe=False)  # 'safe=False' is necessary for serializing non-dict objects


def obtener_secciones(request):
    # Sorted by 'nombre' field in ascending order
    secciones = Seccion.objects.all().order_by('nombre')
    data = [{'id': seccion.id, 'nombre': seccion.nombre} for seccion in secciones]
    return JsonResponse(data, safe=False)  # 'safe=False' is necessary for serializing non-dict objects


def obtener_detalles(request):
    # Sorted by 'nombre' field in ascending order
    detalles = Detalle.objects.all().order_by('rubro')
    data = [{'id': detalle.id, 'rubro': detalle.rubro} for detalle in detalles]
    return JsonResponse(data, safe=False)  # 'safe=False' is necessary for serializing non-dict objects


def get_subsecciones_detalles(request):
    seccion_id = request.GET.get('seccion_id')
    seccion = Seccion.objects.get(id=seccion_id)
    subsecciones = seccion.subseccion_set.all()

    data = {
        'subsecciones': [{
            'id': sub.id,
            'nombre': sub.nombre,
            'detalles': [{
                'id': det.id,
                'rubro': det.rubro,
                'unidad_medida': det.unidad_medida.nombre,
                'cantidad': 0,
                'precio_unitario': det.precio_unitario,
                'precio_total': 0
            } for det in sub.detalle_set.all()]
        } for sub in subsecciones]
    }

    return JsonResponse(data)


def get_detalles(request):
    try:
        subseccion_id = request.GET.get('subseccion_id')
        subseccion = SubSeccion.objects.get(id=subseccion_id)
        detalles = subseccion.detalle_set.all()
        data = {
            'detalles': [{
                'id': det.id,
                'rubro': det.rubro,
                'unidad_medida': det.unidad_medida.nombre,
                'cantidad': 0,
                'precio_unitario': det.precio_unitario,
                'precio_total': 0
            } for det in detalles]
        }

        return JsonResponse(data)
    except ObjectDoesNotExist:
        return JsonResponse({'error': 'SubSeccion not found'}, status=404)
    except Exception as e:
        # Log the exception for debugging
        print(f"Error in get_detalles: {str(e)}")
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


def get_subseccion_name(request):
    subseccion_id = request.GET.get('subseccion_id')
    try:
        subseccion = SubSeccion.objects.get(id=subseccion_id)
        return JsonResponse({'nombre': subseccion.nombre})
    except SubSeccion.DoesNotExist:
        return JsonResponse({'nombre': ''}, status=404)


def get_detalle_data(request):
    det_id = request.GET.get('detalle_id')
    try:
        detalle = Detalle.objects.get(id=det_id)
        data = {
            'id': detalle.id,
            'rubro': detalle.rubro,
            'unidad_medida': detalle.unidad_medida.nombre,
            'cantidad': 0,
            'precio_unitario': detalle.precio_unitario,
            'precio_total': 0
        }
        return JsonResponse(data)
    except ObjectDoesNotExist:
        return JsonResponse({'error': 'Detalle not found'}, status=404)


@csrf_exempt
def crear_seccion(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        subsecciones_ids = request.POST.getlist('subsecciones[]')

        # Create the new Seccion
        new_seccion = Seccion.objects.create(nombre=nombre)

        # Associate the SubSecciones with the new Seccion
        for sub_id in subsecciones_ids:
            subseccion = SubSeccion.objects.get(id=sub_id)
            subseccion.secciones.add(new_seccion)

        return JsonResponse(
            {'seccionId': new_seccion.id, 'status': 'success', 'message': 'Sección creada correctamente.'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@csrf_exempt
def crear_subseccion(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        detalles_ids = request.POST.getlist('detalles[]')
        seccion_id = request.POST.get('id_seccion')
        print('Nombre:', nombre, 'detalles', detalles_ids, 'seccion_id', seccion_id)

        seccion = Seccion.objects.get(id=seccion_id)

        # Create the new Seccion
        new_subseccion = SubSeccion.objects.create(nombre=nombre)
        new_subseccion.secciones.add(seccion)

        # Associate the SubSecciones with the new Seccion
        for det_id in detalles_ids:
            print(det_id)
            detalle = Detalle.objects.get(id=det_id)
            detalle.subsecciones.add(new_subseccion)
        new_subseccion.save()
        return JsonResponse(
            {'subseccionId': new_subseccion.id, 'status': 'success', 'message': 'Subsección creada correctamente.'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@csrf_exempt
def crear_detalle(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        unidadMedida = request.POST.get('unidadMedida')
        cantidad = request.POST.get('cantidad')
        precioUnitario = request.POST.get('precioUnitario')
        precioTotal = request.POST.get('precioTotal')
        subseccion_id = request.POST.get('subseccion')
        print(subseccion_id)
        subseccion = SubSeccion.objects.get(id=subseccion_id)
        unidad = UnidadMedida.objects.get(id=unidadMedida)
        # Create the new Seccion
        detalle = Detalle.objects.create(subseccion=subseccion, rubro=nombre, unidad_medida=unidad, cantidad=cantidad,
                                         precio_unitario=precioUnitario, precio_total=precioTotal)
        data = {
            'id': detalle.id,
            'rubro': detalle.rubro,
            'unidad_medida': detalle.unidad_medida.nombre,
            'cantidad': detalle.cantidad,
            'precio_unitario': detalle.precio_unitario,
            'precio_total': detalle.precio_total
        }
        print(data)
        return JsonResponse(data)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@csrf_exempt
def associate_subseccion(request):
    if request.method == 'POST':
        seccion_id = request.POST.get('seccion_id')
        subseccion_id = request.POST.get('subseccion_id')
        try:
            seccion = Seccion.objects.get(id=seccion_id)
            subseccion = SubSeccion.objects.get(id=subseccion_id)
            subseccion.secciones.add(seccion)
            return JsonResponse({'status': 'success', 'message': 'Subsección asociada correctamente'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request'})


def get_unidad_medida(request):
    unidad_medidas = UnidadMedida.objects.all().values('id', 'descripcion')
    return JsonResponse(list(unidad_medidas), safe=False)

def obtener_marcas_y_unidades(request):
    marcas = Marca.objects.all().values('id', 'nombre')
    unidades_medida = UnidadMedida.objects.all().values('id', 'descripcion')
    data = {
        'marcas': list(marcas),
        'unidades_medida': list(unidades_medida)
    }
    return JsonResponse(data)


def guardar_presupuesto(request):
    if request.method == 'POST':
        try:
            # Carga los datos recibidos en JSON
            data = json.loads(request.body)
            presupuesto_id = data.get('presupuestoId')
            presupuesto = Presupuesto.objects.get(id=presupuesto_id)
            proyecto = presupuesto.proyecto
            obra = Obra.objects.get(proyecto=proyecto)
            obra.plazo = data.get('plazo')
            obra.save()
            presupuesto.subtotal = Decimal(data.get('subtotal', 0))
            presupuesto.iva = Decimal(data.get('iva', 0))
            presupuesto.monto_total = Decimal(data.get('monto_total', 0))
            presupuesto.save()
            # Crea un nuevo objeto Presupuesto, modifica según sea necesario

            # Crea o recupera las secciones
            secciones = [Seccion.objects.get_or_create(id=s_id)[0] for s_id in data['secciones']]

            # Crea o recupera las subsecciones
            subsecciones = []
            for sub in data['subsecciones']:
                obj, created = SubSeccion.objects.get_or_create(id=sub['subseccionId'])
                subsecciones.append(obj)

            # Crea o recupera los detalles
            detalles = []
            for det in data['detalles']:
                obj, created = Detalle.objects.get_or_create(
                    id=det['detalleId'],
                    defaults={
                        'rubro': det['rubro'],
                        'unidad': det['unidad'],
                        'cantidad': det['cantidad'],
                        'precio_unitario': det['precioUnitario'],
                        'precio_total': det['precioTotal'],
                    })
                detalles.append(obj)

            # Crea el ArchivoPresupuesto y asocia las secciones, subsecciones y detalles

            archivo_presupuesto = ArchivoPresupuesto(presupuesto=presupuesto)
            archivo_presupuesto.save()
            archivo_presupuesto.secciones.set(secciones)
            archivo_presupuesto.subsecciones.set(subsecciones)
            archivo_presupuesto.detalles.set(detalles)

            cronograma = Cronograma(archivo_presupuesto=archivo_presupuesto)
            cronograma.save()

            for detalle in detalles:
                detalle_cronograma = DetalleCronograma(cronograma=cronograma,
                                                       detalle=detalle)  # Reemplaza 'fecha' con la fecha que desees
                detalle_cronograma.save()

            return JsonResponse({'status': 'success',
                                 'presupuesto_id': presupuesto_id,  # Aquí devuelves el ID
                                 'message': 'Presupuesto guardado correctamente.'})

        except Exception as e:
            # En caso de error, envía una respuesta con el error
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # Si no es un POST, redirige o envía un error
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def exportar_presupuesto(request, presupuesto_id, archivo_presupuesto_id, tipo):
    print(archivo_presupuesto_id)
    print(presupuesto_id)
    if tipo == 'pdf':
        return exportar_a_pdf(request, archivo_presupuesto_id, presupuesto_id)
    elif tipo == 'excel':
        return exportar_a_excel(request, archivo_presupuesto_id, presupuesto_id)
    else:
        return HttpResponse(status=400)


def exportar_a_pdf(request, archivo_presupuesto_id, presupuesto_id):
    archivo_presupuesto = ArchivoPresupuesto.objects.get(id=archivo_presupuesto_id)
    presupuesto = Presupuesto.objects.get(id=presupuesto_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{presupuesto.proyecto.nombre}-presupuesto.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    story = []

    # Add title
    styles = getSampleStyleSheet()
    title = Paragraph(
        f'Presupuesto del proyecto {presupuesto.proyecto.nombre} para el cliente {presupuesto.proyecto.cliente.nombre}',
        styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))

    # Main data table building
    data = [['Sección', 'Subsección', 'Rubro', 'Unidad Medida', 'Cantidad', 'Precio Unitario', 'Precio Total']]
    for seccion in archivo_presupuesto.secciones.all():
        data.append([seccion.nombre, '', '', '', '', '', ''])
        subsecciones = seccion.subseccion_set.filter(
            id__in=archivo_presupuesto.subsecciones.values_list('id', flat=True))
        for subseccion in subsecciones:
            data.append(['', subseccion.nombre, '', '', '', '', ''])
            detalles = subseccion.detalle_set.filter(id__in=archivo_presupuesto.detalles.values_list('id', flat=True))
            for detalle in detalles:
                data.append([
                    '', '', detalle.rubro, detalle.unidad_medida.nombre,
                    detalle.cantidad, f"{detalle.precio_unitario:,}",
                    f"{detalle.precio_total:,}"
                ])

    # Main data table styling
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    main_table = Table(data)
    main_table.setStyle(table_style)
    story.append(main_table)

    # Spacer between tables
    story.append(Spacer(1, 12))

    # Totals table
    totals_data = [
        ['Subtotal', f"{presupuesto.subtotal:,}"],
        ['IVA', f"{presupuesto.iva:,}"],
        ['Total', f"{presupuesto.monto_total:,}"]
    ]
    totals_table = Table(totals_data, colWidths=[doc.width - 120, 120])
    totals_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(totals_table)

    # Build the PDF
    doc.build(story)

    return response

def exportar_a_excel(request, archivo_presupuesto_id, presupuesto_id):
    archivo_presupuesto = ArchivoPresupuesto.objects.get(id=archivo_presupuesto_id)
    presupuesto = Presupuesto.objects.get(id=presupuesto_id)

    # Create a workbook and grab the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set the title of the sheet
    ws.title = "Presupuesto"

    # Add title
    title = f'Presupuesto del proyecto {presupuesto.proyecto.nombre} para el cliente {presupuesto.proyecto.cliente.nombre}'
    ws.merge_cells('A1:G1')  # Assuming the title spans 7 columns
    ws['A1'] = title
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(bold=True, size=14)

    # Define headers
    headers = ['Sección', 'Subsección', 'Rubro', 'Unidad Medida', 'Cantidad', 'Precio Unitario', 'Precio Total']
    ws.append(headers)

    # Styling headers
    header_fill = PatternFill("solid", fgColor="00AAAAAA")
    for col in range(1, 8):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(border_style="thin"))
        cell.alignment = Alignment(horizontal='center')

    # Adding main data
    row_num = 3
    for seccion in archivo_presupuesto.secciones.all():
        ws.append([seccion.nombre] + [''] * 6)
        subsecciones = seccion.subseccion_set.filter(
            id__in=archivo_presupuesto.subsecciones.values_list('id', flat=True))
        for subseccion in subsecciones:
            ws.append(['', subseccion.nombre] + [''] * 5)
            detalles = subseccion.detalle_set.filter(id__in=archivo_presupuesto.detalles.values_list('id', flat=True))
            for detalle in detalles:
                ws.append([
                    '', '', detalle.rubro, detalle.unidad_medida.nombre,
                    detalle.cantidad, f"{detalle.precio_unitario:,}",
                    f"{detalle.precio_total:,}"
                ])

    # Skip a row before totals
    row_num = ws.max_row + 2

    # Totals table
    ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
    ws.cell(row=row_num, column=5).value = "Subtotal"
    ws.cell(row=row_num, column=7).value = f"{presupuesto.subtotal:,}"

    ws.merge_cells(start_row=row_num + 1, start_column=5, end_row=row_num + 1, end_column=6)
    ws.cell(row=row_num + 1, column=5).value = "IVA"
    ws.cell(row=row_num + 1, column=7).value = f"{presupuesto.iva:,}"

    ws.merge_cells(start_row=row_num + 2, start_column=5, end_row=row_num + 2, end_column=6)
    ws.cell(row=row_num + 2, column=5).value = "Total"
    ws.cell(row=row_num + 2, column=7).value = f"{presupuesto.monto_total:,}"

    # Styling totals
    for row in ws.iter_rows(min_row=row_num, max_row=row_num + 2, min_col=5, max_col=7):
        for cell in row:
            cell.border = Border(top=Side(border_style="thin"),
                                 left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            if cell.column == 7:  # Right align the amount column
                cell.alignment = Alignment(horizontal='right')
            if cell.column == 5:  # Grey background for the label column
                cell.fill = header_fill

        # After filling in all data to the worksheet, iterate over the columns
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]  # Get all cells that are not None in this column
        if column:  # Check if the column is not empty
            for cell in column:
                try:  # Necessary to avoid error on datetime values
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adding a little extra space
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Set the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{presupuesto.proyecto.nombre}-presupuesto.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response


def armar_cronograma(request, obra_id):
    # Obtener la obra y su presupuesto asociado
    obra = get_object_or_404(Obra, id=obra_id)
    proyecto_id = obra.proyecto.id
    proyecto = Proyecto.objects.get(id=proyecto_id)
    presupuesto = Presupuesto.objects.get(proyecto=proyecto)  # Asume que Obra tiene una relación con Presupuesto
    # Obtener el ArchivoPresupuesto asociado
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=presupuesto)
    cronograma = Cronograma.objects.get(archivo_presupuesto=archivo_presupuesto)
    fecha_in = obra.fecha_inicio.strftime('%Y-%m-%d')
    estructura_presupuesto = []
    for seccion in archivo_presupuesto.secciones.all():
        subsecciones_data = []
        subsecciones = seccion.subseccion_set.filter(
            id__in=archivo_presupuesto.subsecciones.values_list('id', flat=True))
        for subseccion in subsecciones:
            detalles = subseccion.detalle_set.filter(id__in=archivo_presupuesto.detalles.values_list('id', flat=True))
            subsecciones_data.append((subseccion, detalles))

        estructura_presupuesto.append((seccion, subsecciones_data))

    # Renderizar una plantilla con los datos
    return render(request, 'pantallas_ing/armar_cronograma.html',
                  {'estructura_presupuesto': estructura_presupuesto, 'obra': obra, 'cronograma': cronograma,
                   'fecha_inicio': fecha_in})


def guardar_cronograma(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            proyecto_id = data.get('proyecto_id')  # Obtén proyecto_id de data

            # Encuentra el proyecto y el cronograma asociado
            proyecto = Proyecto.objects.get(pk=proyecto_id)
            presupuesto = Presupuesto.objects.get(proyecto=proyecto)
            archivo_presupuesto = ArchivoPresupuesto.objects.get(presupuesto=presupuesto)
            cronograma = Cronograma.objects.get(
                archivo_presupuesto=archivo_presupuesto)  # Asumiendo una relación uno a uno entre proyecto y cronograma

            detalles_data = data.get('detalles')  # Obtén la lista de detalles de data
            # Guardar los detalles en la tabla DetalleCronograma
            for detalle_data in detalles_data:
                detalle_id = detalle_data.get('detalleId')
                detalle = Detalle.objects.get(id=detalle_id)
                detalle_cronograma = DetalleCronograma.objects.get(detalle=detalle, cronograma=cronograma)
                fecha_seleccionada = detalle_data.get('fechaSeleccionada')
                detalle_cronograma.fecha_programada = fecha_seleccionada
                detalle_cronograma.save()

            return JsonResponse({'status': 'success', 'message': 'Cronograma guardado con éxito'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})


def vista_redireccion(request, campo, valor):
    if campo == 'cliente':
        cliente = Cliente.objects.get(nombre=valor)
        presupuestos = Presupuesto.objects.filter(proyecto__cliente_id=cliente.id, encargado=request.user)
    elif campo == 'estadoPresupuesto':
        if valor == 'Aprobado':  # Verificar si el valor es 'Aprobado'
            presupuestos = Presupuesto.objects.filter(estado='A', encargado=request.user)  # Filtrar por estado 'A'
        elif valor == 'En elaboración':
            presupuestos = Presupuesto.objects.filter(estado='E', encargado=request.user)  # Filtrar por estado 'A'
        elif valor == 'Enviado':
            presupuestos = Presupuesto.objects.filter(estado='S', encargado=request.user)  # Filtrar por estado 'A'
        else:
            presupuestos = Presupuesto.objects.filter(encargado=request.user)
    else:
        presupuestos = Presupuesto.objects.filter(encargado=request.user)

    return render(request, 'pantallas_ing/ver_presupuestos.html', {'presupuestos': presupuestos})


def vista_redireccion_obra(request, campo, valor):
    if campo == 'cliente':
        cliente = Cliente.objects.get(nombre=valor)
        obras = Obra.objects.filter(proyecto__cliente_id=cliente.id, encargado=request.user)
    elif campo == 'estadoObra':
        if valor == 'No iniciada':  # Verificar si el valor es 'Aprobado'
            obras = Obra.objects.filter(estado='NI', encargado=request.user)  # Filtrar por estado 'A'
        elif valor == 'En ejecución':
            obras = Obra.objects.filter(estado='E', encargado=request.user)  # Filtrar por estado 'A'
        elif valor == 'Finalizada':
            obras = Obra.objects.filter(estado='F', encargado=request.user)  # Filtrar por estado 'A'
        else:
            obras = Obra.objects.filter(encargado=request.user)
    else:
        obras = Obra.objects.filter(encargado=request.user)
    # Iterar sobre las obras y agregar parámetros adicionales
    obras_con_parametros = []
    for obra in obras:
        # Obtener el cronograma asociado a la obra
        cronograma = Cronograma.objects.get(archivo_presupuesto__presupuesto__proyecto=obra.proyecto)
        # Verificar si existe un cronograma y si al menos un detalle tiene fecha
        if cronograma and cronograma.detalles_cronograma.filter(fecha_programada__isnull=False).exists():
            cronograma_existente = True
        else:
            cronograma_existente = False
        obra_parametros = {
            'obra': obra,
            'cronograma_existe': cronograma_existente,
            'cronograma': cronograma
        }
        obras_con_parametros.append(obra_parametros)

    return render(request, 'pantallas_ing/ver_obras.html', {'obras': obras_con_parametros})


def ver_proyectos_filtrados(request):
    campo = request.GET.get('campo')
    valor = request.GET.get('valor')
    # Generar la URL de redirección a 'ver_presupuestos' con los parámetros deseados
    redirect_url = f'/vista_redireccion/{campo}/{valor}/'

    # Devolver una respuesta JSON con la URL de redirección
    return JsonResponse({'redirect_url': redirect_url})


def ver_presupuestos_filtrados(request):
    campo = request.GET.get('campo')
    valor = request.GET.get('valor')
    # Generar la URL de redirección a 'ver_presupuestos' con los parámetros deseados
    redirect_url = f'/vista_redireccion/{campo}/{valor}/'

    # Devolver una respuesta JSON con la URL de redirección
    return JsonResponse({'redirect_url': redirect_url})


def ver_obras_filtrados(request):
    campo = request.GET.get('campo')
    valor = request.GET.get('valor')
    # Generar la URL de redirección a 'ver_presupuestos' con los parámetros deseados
    redirect_url = f'/vista_redireccion_obra/{campo}/{valor}/'

    # Devolver una respuesta JSON con la URL de redirección
    return JsonResponse({'redirect_url': redirect_url})


def buscar_materiales(request):
    search_query = request.GET.get('search', '')

    # Realiza una búsqueda en la base de datos utilizando el modelo Material
    # Puedes ajustar esta consulta según tus necesidades específicas
    resultados = Material.objects.filter(Q(nombre__icontains=search_query))

    # Crea una lista de diccionarios con los resultados
    resultados_list = [{'id': material.id, 'nombre': material.nombre} for material in resultados]

    # Devuelve los resultados en formato JSON
    return JsonResponse(resultados_list, safe=False)


def ver_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    materiales_pedido = MaterialPedido.objects.filter(
        pedido=pedido)  # Obtiene los materiales relacionados con el pedido
    return render(request, 'pantallas_ing/ver_pedido.html', {'pedido': pedido, 'materiales_pedido': materiales_pedido})

def ver_pedido_deposito(request, pedido_id):
    # Obtener el pedido
    pedido = Pedido.objects.get(pk=pedido_id)

    # Obtener los materiales requeridos para el pedido
    materiales_requeridos = defaultdict(int)
    for material_pedido in pedido.materialpedido_set.all():
        cantidad_requerida = material_pedido.cantidad
        materiales_requeridos[material_pedido.material] += cantidad_requerida

    # Verificar si hay suficiente stock para cada material
    materiales_disponibles = []
    materiales_a_comprar = []
    for material, cantidad_requerida in materiales_requeridos.items():
        stock_disponible = material.unidades_stock - material.minimo
        if stock_disponible >= cantidad_requerida:
            materiales_disponibles.append(material)
        else:
            materiales_a_comprar.append(material)
    context = {
        'pedido': pedido,
        'materiales_requeridos': materiales_requeridos,
        'materiales_disponibles': materiales_disponibles,
        'materiales_a_comprar': materiales_a_comprar
    }
    return render(request, 'pantallas_deposito/ver_pedido_dep.html', context)

def entregar_pedido(request, pedido_id):
    if request.method == "POST":
        try:
            pedido = Pedido.objects.get(pk=pedido_id)
            materiales_pedido = MaterialPedido.objects.filter(pedido=pedido)

            # Actualizar el stock de cada material en el pedido
            for material_pedido in materiales_pedido:
                material = material_pedido.material
                if material.unidades_stock >= material_pedido.cantidad:
                    material.unidades_stock -= material_pedido.cantidad
                else:
                    return JsonResponse({'status': 'error', 'message': f'Stock insuficiente para el material {material.nombre}'}, status=400)
                material.save()

            # Actualizar el estado del pedido
            pedido.estado = 'E'
            pedido.fecha_entrega = date.today()
            pedido.save()

            return redirect('ver_pedidos_dep')
        except Pedido.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Pedido no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def ver_pedido_adm(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    materiales_pedido = MaterialPedido.objects.filter(
        pedido=pedido)  # Obtiene los materiales relacionados con el pedido
    obra = pedido.obra
    return render(request, 'pantallas_adm/ver_pedido_adm.html',
                  {'pedido': pedido, 'materiales_pedido': materiales_pedido, 'obra': obra})


def ver_cronograma(request, obra_id, cronograma_id):
    obra = get_object_or_404(Obra, id=obra_id)
    cronograma = get_object_or_404(Cronograma, id=cronograma_id)
    detalles_cronograma = cronograma.detalles_cronograma.all()

    # Construir la estructura jerárquica de secciones y subsecciones
    estructura_presupuesto = defaultdict(lambda: defaultdict(list))

    # Verificar que realmente estamos obteniendo secciones y subsecciones
    for seccion in Seccion.objects.filter(archivopresupuesto=cronograma.archivo_presupuesto):
        for subseccion in SubSeccion.objects.filter(secciones=seccion,
                                                    archivopresupuesto=cronograma.archivo_presupuesto):
            print("Sección:", seccion.nombre, "Subsección:", subseccion.nombre)  # Depuración
            for detalle in Detalle.objects.filter(subseccion=subseccion,
                                                  archivopresupuesto=cronograma.archivo_presupuesto):
                detalle_cron = DetalleCronograma.objects.get(detalle=detalle, cronograma=cronograma)
                fecha_programada = detalle_cron.fecha_programada
                fecha_culminacion = detalle_cron.fecha_culminacion
                realizado = detalle_cron.realizado
                estructura_presupuesto[seccion][subseccion].append(
                    (detalle, fecha_programada, fecha_culminacion, realizado))

    # Convertir cada defaultdict interno a un diccionario regular
    estructura_presupuesto_regular = {seccion: dict(subsecciones) for seccion, subsecciones in
                                      estructura_presupuesto.items()}

    context = {
        'obra': obra,
        'cronograma': cronograma,
        'estructura_presupuesto': estructura_presupuesto_regular,
        'hoy': date.today(),  # Agrega la fecha actual al contexto
    }
    return render(request, 'pantallas_ing/ver_cronograma.html', context)


def ver_cronograma_adm(request, obra_id):
    obra = get_object_or_404(Obra, id=obra_id)
    proyecto = obra.proyecto
    presupuesto = Presupuesto.objects.get(proyecto=proyecto)
    archivo_presupuesto = ArchivoPresupuesto.objects.get(presupuesto=presupuesto)
    cronograma = Cronograma.objects.get(archivo_presupuesto=archivo_presupuesto)
    detalles_cronograma = cronograma.detalles_cronograma.all()
    detalles_fechas_n = cronograma.detalles_cronograma.filter(fecha_programada__isnull=True)
    if detalles_fechas_n.exists():
        messages.warning(request, "No se ha creado el cronograma aún")
        return redirect('ver_obras_adm')  # Ajusta el nombre de la vista según tu configuración de URL

    # Crear un diccionario que mapea los detalles a sus fechas en el cronograma
    detalles_fechas = {detalle_crono.detalle: detalle_crono.fecha_programada for detalle_crono in detalles_cronograma}

    # Construir la estructura jerárquica de secciones y subsecciones
    estructura_presupuesto = defaultdict(lambda: defaultdict(list))

    # Verificar que realmente estamos obteniendo secciones y subsecciones
    for seccion in Seccion.objects.filter(archivopresupuesto=cronograma.archivo_presupuesto):
        for subseccion in SubSeccion.objects.filter(secciones=seccion,
                                                    archivopresupuesto=cronograma.archivo_presupuesto):
            print("Sección:", seccion.nombre, "Subsección:", subseccion.nombre)  # Depuración
            for detalle in Detalle.objects.filter(subseccion=subseccion,
                                                  archivopresupuesto=cronograma.archivo_presupuesto):
                detalle_cron = DetalleCronograma.objects.get(detalle=detalle, cronograma=cronograma)
                fecha_programada = detalle_cron.fecha_programada
                fecha_culminacion = detalle_cron.fecha_culminacion
                realizado = detalle_cron.realizado
                estructura_presupuesto[seccion][subseccion].append(
                    (detalle, fecha_programada, fecha_culminacion, realizado))

    # Convertir cada defaultdict interno a un diccionario regular
    estructura_presupuesto_regular = {seccion: dict(subsecciones) for seccion, subsecciones in
                                      estructura_presupuesto.items()}

    context = {
        'obra': obra,
        'cronograma': cronograma,
        'estructura_presupuesto': estructura_presupuesto_regular,
    }
    return render(request, 'pantallas_adm/ver_cronograma_adm.html', context)


def presupuestos_pendientes(request):
    # Obtener los presupuestos que están en estado 'E' y pertenecen al usuario actual
    presupuestos = Presupuesto.objects.filter(encargado=request.user, estado='E')

    # Agregar una anotación para verificar si existe al menos un DetalleCronograma asociado
    presupuestos = presupuestos.annotate(
        tiene_detalle_cronograma=Exists(
            DetalleCronograma.objects.filter(
                cronograma__archivo_presupuesto__presupuesto=OuterRef('pk')
            )
        )
    )
    presupuestos = presupuestos.filter(
        (Q(tiene_detalle_cronograma=False))
    )

    data = [
        {
            'id': presupuesto.id,
            'proyecto__nombre': presupuesto.proyecto.nombre,
            'crear_presupuesto_url': reverse('crear_presupuesto', args=[presupuesto.id])
        }
        for presupuesto in presupuestos
    ]
    return JsonResponse({'presupuestos': data})


def obras_pendientes(request):
    obras = Obra.objects.filter(encargado=request.user, estado__in=['E', 'NI'])
    data = []
    for obra in obras:
        acciones = []
        cronograma = Cronograma.objects.get(archivo_presupuesto__presupuesto__proyecto=obra.proyecto)
        if cronograma and cronograma.detalles_cronograma.filter(fecha__isnull=False).exists():
            cronograma_existente = True
        else:
            cronograma_existente = False
        if obra.estado == 'NI':
            if not obra.fecha_inicio:
                acciones.append({'nombre': 'Agendar Inicio',
                                 'html': '<button class="btn btn-primary agendar-inicio-obra" data-id="' + str(
                                     obra.id) + '" style="margin-right: 5px;">Agendar Inicio</button>'})
            elif not cronograma_existente:
                acciones.append({'nombre': 'Armar Cronograma', 'url': reverse('armar_cronograma', args=[obra.id])})
        if acciones:
            data.append({'id': obra.id, 'proyecto__nombre': obra.proyecto.nombre, 'acciones': acciones})
        print(data)
    return JsonResponse({'obras': data})


def proximas_actividades(request):
    fecha_limite = timezone.now().date() + timedelta(days=7)
    actividades = DetalleCronograma.objects.filter(
        fecha_programada__lte=fecha_limite,
        fecha_culminacion=None,
        realizado=False,
        fecha_programada__gte=timezone.now().date()
    ).select_related('detalle', 'cronograma', 'cronograma__archivo_presupuesto',
                     'cronograma__archivo_presupuesto__presupuesto',
                     'cronograma__archivo_presupuesto__presupuesto__proyecto')

    data = [
        {
            'proyecto_nombre': actividad.cronograma.archivo_presupuesto.presupuesto.proyecto.nombre,
            'rubro': actividad.detalle.rubro,
            'fecha_programada': actividad.fecha_programada.strftime('%d-%m-%Y'),
            'fecha_programada': actividad.fecha_culminacion.strftime('%d-%m-%Y'),
            'detalle_id': actividad.detalle.id,
            'cronograma_id': actividad.cronograma.id
        }
        for actividad in actividades
    ]
    return JsonResponse({'actividades': data})


@csrf_exempt
def marcar_como_realizado(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        cronograma_id = data.get('cronograma_id')
        detalle_id = data.get('detalle_id')
        print(detalle_id)
        try:
            detalle_cronograma = DetalleCronograma.objects.get(
                cronograma_id=cronograma_id, detalle_id=detalle_id
            )
            detalle_cronograma.realizado = True
            detalle_cronograma.fecha_culminacion = timezone.now().date()
            detalle_cronograma.save()
            return JsonResponse({'status': 'success'})
        except DetalleCronograma.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'DetalleCronograma no encontrado'})

    return JsonResponse({'status': 'error', 'message': 'Solicitud inválida'})


def presupuestos_en_elaboracion(request):
    presupuestos = Presupuesto.objects.filter(estado='E').select_related('proyecto__cliente')

    data = [
        {
            'proyecto': presupuesto.proyecto.nombre,
            'cliente': presupuesto.proyecto.cliente.nombre,  # Ajusta este campo según tu modelo Cliente
            'encargado': f'{presupuesto.encargado.first_name} {presupuesto.encargado.last_name}' if presupuesto.encargado else 'No asignado',
        }
        for presupuesto in presupuestos
    ]

    return JsonResponse({'presupuestos': data})


def presupuestos_enviados_sin_anticipo(request):
    presupuestos = Presupuesto.objects.filter(
        estado='S',
        anticipo=False
    )
    print(len(presupuestos))
    data = [
        {
            'proyecto': presupuesto.proyecto.nombre,
            'cliente': presupuesto.proyecto.cliente.nombre,
            'encargado': f'{presupuesto.encargado.first_name} {presupuesto.encargado.last_name}' if presupuesto.encargado else 'No asignado',
        }
        for presupuesto in presupuestos
    ]

    return JsonResponse({'presupuestos': data})


def obras_pendientes_de_asignacion(request):
    obras = Obra.objects.filter(
        encargado__isnull=True,
        proyecto__presupuesto__estado='A',
        proyecto__presupuesto__anticipo=True
    ).select_related('proyecto__cliente')

    data = [
        {
            'proyecto': obra.proyecto.nombre,
            'cliente': obra.proyecto.cliente.nombre,  # Asegúrate de que esta línea corresponda con tu modelo Cliente
            'acciones': f'<a href="/obra/{obra.id}/detalle/">Ver Detalles</a>'  # Asegúrate de tener esta URL
        }
        for obra in obras
    ]

    return JsonResponse({'obras': data})


def ver_certificados_ing(request):
    # Obtiene todas las obras donde el usuario es el encargado
    obras = Obra.objects.filter(encargado=request.user)
    certificados = []

    for obra in obras:
        # Obtiene el proyecto relacionado con la obra
        proyecto = obra.proyecto
        # Obtiene todos los presupuestos relacionados con el proyecto
        presupuestos = Presupuesto.objects.filter(proyecto=proyecto)
        # Para cada presupuesto, obtiene los certificados relacionados y los añade a la lista
        for presupuesto in presupuestos:
            for certificado in Certificado.objects.filter(presupuesto=presupuesto):
                certificados.append(certificado)

    # Renderiza la respuesta pasando la lista de certificados al contexto
    return render(request, 'pantallas_ing/ver_certificados.html', {
        'title': 'Certificados',
        'certificados': certificados
    })


def ver_certificados_adm(request):
    certificados = Certificado.objects.all()
    ingenieros = User.objects.filter(groups__name='INGENIERO')
    clientes = Cliente.objects.all()
    proyectos = Proyecto.objects.all()

    query = request.GET.get('q')
    ingeniero_id = request.GET.get('ingeniero')
    cliente_id = request.GET.get('cliente')
    proyecto_id = request.GET.get('proyecto')

    if query:
        certificados = certificados.filter(
            Q(presupuesto__proyecto__nombre__icontains=query) |
            Q(presupuesto__proyecto__cliente__nombre__icontains=query)
        )
    if ingeniero_id:
        certificados = certificados.filter(presupuesto__proyecto__ingeniero_id=ingeniero_id)
    if cliente_id:
        certificados = certificados.filter(presupuesto__proyecto__cliente_id=cliente_id)
    if proyecto_id:
        certificados = certificados.filter(presupuesto__proyecto_id=proyecto_id)

    print(ingenieros)
    return render(request, 'pantallas_adm/ver_certificados_adm.html', {
        'title': 'Certificados',
        'certificados': certificados,
        'ingenieros': ingenieros,
        'clientes': clientes,
        'proyectos': proyectos,
    })


def ver_certificado_adm(request, pk):
    archivo_certificado = get_object_or_404(ArchivoCertificado, certificado=pk)
    certificado = Certificado.objects.get(id=pk)
    # Estructura para mantener la jerarquía
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=certificado.presupuesto)
    estructura_certificado = {}
    subtotal = certificado.subtotal
    iva = certificado.iva
    monto_total = certificado.monto_total
    cliente = certificado.presupuesto.proyecto.cliente.nombre
    proyecto = certificado.presupuesto.proyecto.nombre
    # Iterar sobre los detalles asociados con el archivo de certificado
    for detalle in archivo_certificado.detalles.all():
        subseccion = detalle.subseccion
        secciones_subseccion = set(subseccion.secciones.all())
        secciones_presupuesto = set(archivo_presupuesto.secciones.all())
        secciones_comunes = secciones_subseccion.intersection(secciones_presupuesto)
        seccion = next(iter(secciones_comunes), None)
        if seccion not in estructura_certificado:
            estructura_certificado[seccion] = {}

        if subseccion not in estructura_certificado[seccion]:
            estructura_certificado[seccion][subseccion] = []

        estructura_certificado[seccion][subseccion].append(detalle)

    context = {
        'estructura_certificado': estructura_certificado,
        'subtotal': subtotal,
        'monto_total': monto_total,
        'iva': iva,
        'proyecto': proyecto,
        'cliente': cliente,
        'certificado_id': certificado.id,
        'archivo_certificado_id': archivo_certificado.id,
        'comprobante': certificado.comprobante_pago
    }

    return render(request, 'pantallas_adm/ver_certificado_adm.html', context)


def ver_certificado_gerente(request, pk):
    archivo_certificado = get_object_or_404(ArchivoCertificado, certificado=pk)
    certificado = Certificado.objects.get(id=pk)
    # Estructura para mantener la jerarquía
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=certificado.presupuesto)
    estructura_certificado = {}
    subtotal = certificado.subtotal
    iva = certificado.iva
    monto_total = certificado.monto_total
    cliente = certificado.presupuesto.proyecto.cliente.nombre
    proyecto = certificado.presupuesto.proyecto
    proyecto_nombre = proyecto.nombre
    obra = Obra.objects.get(proyecto=proyecto)
    # Iterar sobre los detalles asociados con el archivo de certificado
    for detalle in archivo_certificado.detalles.all():
        subseccion = detalle.subseccion
        secciones_subseccion = set(subseccion.secciones.all())
        secciones_presupuesto = set(archivo_presupuesto.secciones.all())
        secciones_comunes = secciones_subseccion.intersection(secciones_presupuesto)
        seccion = next(iter(secciones_comunes), None)
        if seccion not in estructura_certificado:
            estructura_certificado[seccion] = {}

        if subseccion not in estructura_certificado[seccion]:
            estructura_certificado[seccion][subseccion] = []

        estructura_certificado[seccion][subseccion].append(detalle)

    context = {
        'estructura_certificado': estructura_certificado,
        'subtotal': subtotal,
        'monto_total': monto_total,
        'iva': iva,
        'proyecto_nombre': proyecto_nombre,
        'cliente': cliente,
        'certificado_id': certificado.id,
        'archivo_certificado_id': archivo_certificado.id,
        'comprobante': certificado.comprobante_pago,
        'obra': obra
    }

    return render(request, 'pantallas_gerente/ver_certificado_gerente.html', context)


def elaborar_certificado(request):
    obras = Obra.objects.filter(encargado=request.user, estado='E')
    context = {
        'obras': obras,
    }

    return render(request, 'pantallas_ing/elaborar_certificado.html', context)


def cargar_detalles_certificado(request):
    obra_id = request.GET.get('obra_id')
    obra = get_object_or_404(Obra, id=obra_id)
    proyecto_id = obra.proyecto.id

    # Obtener los IDs de los detalles que ya han sido certificados mediante ArchivoCertificado
    detalles_certificados_ids = Detalle.objects.filter(
        archivocertificado__certificado__presupuesto__proyecto__id=proyecto_id
    ).values_list('id', flat=True)

    # Filtrar los detalles del cronograma que no han sido certificados
    detalles = DetalleCronograma.objects.filter(
        cronograma__archivo_presupuesto__presupuesto__proyecto__id=proyecto_id,
        realizado=True
    ).exclude(
        detalle_id__in=detalles_certificados_ids
    ).values(
        'id', 'detalle__rubro', 'fecha_programada', 'fecha_culminacion',
        'detalle__cantidad', 'detalle__precio_unitario', 'detalle__precio_total'
    )

    return JsonResponse(list(detalles), safe=False)


@csrf_exempt
def guardar_certificado(request):
    if request.method == 'POST':
        ingeniero = request.user
        data = json.loads(request.body)
        ids = data.get('ids')
        obra_id = data.get('obra_id')

        try:
            obra = Obra.objects.get(id=obra_id)
            certificado = Certificado.objects.create(ingeniero=ingeniero, presupuesto=obra.proyecto.presupuesto, monto_total=0,subtotal=0, iva=0)  # Asocia directamente con el presupuesto de la obra
            archivo_certificado = ArchivoCertificado.objects.create(certificado=certificado)  # Crea un archivo de certificado para asociar detalles
            subtotal = Decimal('0.00')
            tasa_iva = Decimal('0.10')  # 10% de IVA como Decimal

            for identificador in ids:
                detalle_cronograma = DetalleCronograma.objects.get(id=identificador)
                archivo_certificado.detalles.add(detalle_cronograma.detalle)
                subtotal += Decimal(str(detalle_cronograma.detalle.precio_total))

            # Calcula el IVA como el 10% del subtotal (ajusta la tasa de IVA según corresponda)
            iva = subtotal * tasa_iva

            # Calcula el monto total sumando el subtotal y el IVA
            monto_total = subtotal + iva

            certificado.subtotal = subtotal
            certificado.iva = iva
            certificado.monto_total = monto_total
            certificado.save()

            return JsonResponse({"status": "success"})
        except Obra.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Obra no encontrada"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error"}, status=400)


@csrf_exempt
def marcar_certificado_enviado(request, certificado_id):
    try:
        # Convertir el body de la solicitud de JSON a un diccionario de Python
        data = json.loads(request.body)

        # Encontrar el certificado con el ID proporcionado
        certificado = Certificado.objects.get(id=certificado_id)

        # Cambiar el estado y la fecha de envío del certificado
        certificado.estado = 'S'
        certificado.fecha_envio = date.today()

        # Guardar los cambios en la base de datos
        certificado.save()

        # Devolver una respuesta de éxito
        return JsonResponse({"success": True, "message": "Certificado marcado como enviado."})

    except Certificado.DoesNotExist:
        return JsonResponse({"success": False, "message": "Certificado no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


@csrf_exempt
def registrar_pago_certificado(request, certificado_id):
    try:
        if request.method == 'POST':
            # Obtener el certificado
            certificado = Certificado.objects.get(id=certificado_id)

            # Comprobar si se ha subido un comprobante de pago
            comprobante = request.FILES.get('comprobantePago')
            if not comprobante:
                return JsonResponse({'success': False, 'message': 'No se ha proporcionado un comprobante de pago.'})

            # Limpiar el nombre del archivo y reemplazar espacios con guiones bajos
            comprobante_name = get_valid_filename(comprobante.name)
            cleaned_filename = comprobante_name.replace(' ', '_')

            # Guardar el comprobante de pago
            fs = FileSystemStorage()
            filename = fs.save(cleaned_filename, comprobante)
            certificado.comprobante_pago = fs.url(filename)

            # Actualizar la fecha de pago y el estado
            certificado.fecha_pago = timezone.now().date()
            certificado.estado = 'A'
            certificado.save()

            return JsonResponse({'success': True, 'message': 'Pago registrado exitosamente.'})
        else:
            return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

    except Certificado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Certificado no encontrado.'}, status=404)
    except Exception as e:
        logger.error(f"Error registrando pago: {str(e)}")  # Asegúrate de tener configurado un logger
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def ver_archivo_certificado(request, pk):
    archivo_certificado = get_object_or_404(ArchivoCertificado, certificado=pk)
    certificado = Certificado.objects.get(id=pk)
    # Estructura para mantener la jerarquía
    archivo_presupuesto = get_object_or_404(ArchivoPresupuesto, presupuesto=certificado.presupuesto)
    estructura_certificado = {}
    subtotal = certificado.subtotal
    iva = certificado.iva
    monto_total = certificado.monto_total
    cliente = certificado.presupuesto.proyecto.cliente.nombre
    proyecto = certificado.presupuesto.proyecto.nombre
    # Iterar sobre los detalles asociados con el archivo de certificado
    for detalle in archivo_certificado.detalles.all():
        subseccion = detalle.subseccion
        secciones_subseccion = set(subseccion.secciones.all())
        secciones_presupuesto = set(archivo_presupuesto.secciones.all())
        secciones_comunes = secciones_subseccion.intersection(secciones_presupuesto)
        seccion = next(iter(secciones_comunes), None)
        if seccion not in estructura_certificado:
            estructura_certificado[seccion] = {}

        if subseccion not in estructura_certificado[seccion]:
            estructura_certificado[seccion][subseccion] = []

        estructura_certificado[seccion][subseccion].append(detalle)

    context = {
        'estructura_certificado': estructura_certificado,
        'subtotal': subtotal,
        'monto_total': monto_total,
        'iva': iva,
        'proyecto': proyecto,
        'cliente': cliente,
        'certificado_id': certificado.id,
        'archivo_certificado_id': archivo_certificado.id
    }

    return render(request, 'pantallas_ing/ver_certificado.html', context)


@csrf_exempt
def aceptar_devolucion(request, devolucion_id):
    if request.method == "POST":
        try:
            devolucion = Devolucion.objects.get(pk=devolucion_id)
            materiales_devueltos = MaterialDevuelto.objects.filter(devolucion=devolucion)

            # Actualizar el stock de cada material devuelto
            for material_devuelto in materiales_devueltos:
                material = material_devuelto.material
                material.unidades_stock += material_devuelto.cantidad
                material.save()

            # Actualizar el estado de la devolución
            devolucion.estado = 'D'
            devolucion.fecha_devolucion = date.today()
            devolucion.save()

            return HttpResponseRedirect(reverse('ver_devoluciones_dep'))
        except Devolucion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devolución no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@csrf_exempt
def rechazar_devolucion(request, devolucion_id):
    if request.method == "POST":
        try:
            devolucion = Devolucion.objects.get(pk=devolucion_id)
            devolucion.estado = 'R'
            observaciones = request.POST.get('observaciones')
            print("Observaciones enviadas:", observaciones)  # Imprimir las observaciones enviadas
            devolucion.observaciones = observaciones
            devolucion.save()
            print("Observaciones guardadas:", devolucion.observaciones)  # Imprimir las observaciones guardadas
            return HttpResponseRedirect(reverse('ver_devoluciones_dep'))
        except Devolucion.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Devolución no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

def ver_devolucion(request, devolucion_id):
    devolucion = get_object_or_404(Devolucion, id=devolucion_id)
    context = {
        'devolucion': devolucion
    }
    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_devolucion_dep.html'
    elif request.user.groups.filter(name='GERENTE').exists():
        template_name = 'pantallas_gerente/ver_devolucion_gerente.html'
    else:
        template_name = 'pantallas_ing/ver_devolucion.html'
    return render(request, template_name, context)


def ver_devoluciones(request):
    estado_seleccionado = request.GET.get('estado')
    obra_seleccionada = request.GET.get('obra')
    encargado_seleccionado = request.GET.get('encargado')

    # Inicializar el queryset
    devoluciones = Devolucion.objects.all().annotate(
        is_pending=Case(
            When(estado='P', then=Value(0)),
            When(estado='D', then=Value(1)),
            When(estado='R', then=Value(2)),
            output_field=IntegerField(),
        )
    )

    # Aplicar filtros
    if estado_seleccionado:
        devoluciones = devoluciones.filter(estado=estado_seleccionado)
    if obra_seleccionada:
        try:
            obra_seleccionada = int(obra_seleccionada)
            devoluciones = devoluciones.filter(obra_id=obra_seleccionada)
        except ValueError:
            pass
    if encargado_seleccionado:
        try:
            encargado_seleccionado = int(encargado_seleccionado)
            devoluciones = devoluciones.filter(obra__encargado_id=encargado_seleccionado)
        except ValueError:
            pass

    # Ordenar los resultados
    devoluciones = devoluciones.order_by('is_pending', '-fecha_solicitud')

    # Determinar qué template renderizar en función del rol del usuario
    if request.user.groups.filter(name='ENCARGADO_DEPOSITO').exists():
        template_name = 'pantallas_deposito/ver_devoluciones_dep.html'
    else:
        template_name = 'pantallas_ing/ver_devoluciones.html'

    # Obtener todas las obras y encargados para los filtros
    obras = Obra.objects.all().values('id', 'proyecto__nombre')
    encargados = User.objects.filter(groups__name='INGENIERO').values('id', 'first_name', 'last_name')

    context = {
        'devoluciones': devoluciones,
        'estado_seleccionado': estado_seleccionado,
        'obra_seleccionada': obra_seleccionada,
        'encargado_seleccionado': encargado_seleccionado,
        'obras': list(obras),  # Convertimos a lista de diccionarios
        'encargados': list(encargados),  # Convertimos a lista de diccionarios
    }

    return render(request, template_name, context)

def devoluciones_pedidos_pendientes(request):
    devoluciones_pendientes = Devolucion.objects.filter(estado='P').values('obra__nombre')
    pedidos_pendientes = Pedido.objects.filter(estado='P').values('id')

    data = {
        'devoluciones': list(devoluciones_pendientes),
        'pedidos': list(pedidos_pendientes)
    }

    return JsonResponse(data)


def ver_inventario(request):
    materiales = Material.objects.all().order_by('nombre')
    form_buscar = BuscadorMaterialForm()

    # Configuración de la paginación
    paginator = Paginator(materiales, 10)  # Muestra 10 materiales por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lógica para eliminar el material
    material_id = request.GET.get('eliminar_material_id')
    if material_id:
        material = get_object_or_404(Material, id=material_id)

        # Verificamos si el material está siendo utilizado en pedidos
        pedidos_asociados = Pedido.objects.filter(materiales=material)

        # Verificamos si las obras asociadas a esos pedidos están activas
        obras_activas_asociadas = Obra.objects.filter(pedido__in=pedidos_asociados, estado='E')

        if obras_activas_asociadas.exists():
            # Si hay obras activas que hacen referencia al material, mostramos un mensaje de error
            messages.error(request, 'No se puede eliminar el material porque está siendo utilizado en obras activas.')
        else:
            material.delete()
            messages.success(request, 'Material eliminado exitosamente.')

    if request.method == 'GET':
        form_buscar = BuscadorMaterialForm(request.GET)
        if form_buscar.is_valid():
            termino_busqueda = form_buscar.cleaned_data['termino_busqueda']
            if termino_busqueda:
                materiales = materiales.filter(nombre__icontains=termino_busqueda)
                paginator = Paginator(materiales, 10)  # Actualiza la paginación con los resultados filtrados
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)

    return render(request, 'pantallas_deposito/ver_inventario.html',
                  {'materiales': materiales, 'form_buscar': form_buscar, 'page_obj': page_obj})


def get_obras_activas(request):
    cliente_id = request.GET.get('cliente')
    ingeniero_id = request.GET.get('ingeniero')
    page_number = request.GET.get('page', 1)

    obras_activas = Obra.objects.filter(estado='E').select_related('proyecto', 'encargado', 'proyecto__cliente')
    print(obras_activas)
    if cliente_id:
        obras_activas = obras_activas.filter(proyecto__cliente_id=cliente_id)
    if ingeniero_id:
        obras_activas = obras_activas.filter(encargado_id=ingeniero_id)

    paginator = Paginator(obras_activas, 10)
    page_obj = paginator.get_page(page_number)

    obras_data = []
    for obra in page_obj:
        obras_data.append({
            'id': obra.id,
            'proyecto': {
                'nombre': obra.proyecto.nombre,
                'cliente': {
                    'nombre': obra.proyecto.cliente.nombre
                }
            },
            'encargado': {
                'get_full_name': obra.encargado.get_full_name()
            }
        })

    data = {
        'obras': obras_data,
        'page_obj': {
            'number': page_obj.number,
            'paginator': {
                'num_pages': paginator.num_pages,
            },
            'has_previous': page_obj.has_previous(),
            'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'has_next': page_obj.has_next(),
            'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        }
    }

    return JsonResponse(data)


def get_ingenieros_clientes_obras_activas():
    # Filtrar obras activas
    obras_activas = Obra.objects.filter(estado='E').select_related('proyecto', 'encargado', 'proyecto__cliente')

    # Obtener clientes únicos de las obras activas
    clientes_ids = obras_activas.values_list('proyecto__cliente_id', flat=True).distinct()
    clientes = Cliente.objects.filter(id__in=clientes_ids)

    # Obtener ingenieros encargados únicos de las obras activas
    ingenieros_ids = obras_activas.values_list('encargado_id', flat=True).distinct()
    ingenieros_encargados = User.objects.filter(id__in=ingenieros_ids)

    return clientes, ingenieros_encargados


def generar_grafico_materiales(hechos_materiales):
    df = pd.DataFrame(list(hechos_materiales.values('material_id', 'material_nombre', 'cantidad_total')))

    fig = go.Figure()

    for material in df['material_nombre'].unique():
        df_material = df[df['material_nombre'] == material]
        fig.add_trace(go.Bar(
            x=df_material['material_nombre'],
            y=df_material['cantidad_total'],
            name=material
        ))

    # Convertir el gráfico a JSON
    graph_json = json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)
    return graph_json


def generar_grafico_certificados(hechos_certificados):
    # Convertir los datos a un DataFrame
    df = pd.DataFrame(list(hechos_certificados.values('id','fecha_envio', 'monto_total')))

    # Filtrar filas con valores nulos
    df = df.dropna(subset=['fecha_envio', 'monto_total'])

    # Asegurarse de que las fechas están en el formato correcto
    df['fecha_envio'] = pd.to_datetime(df['fecha_envio'])

    # Crear el gráfico de líneas
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df['fecha_envio'],
        y=df['monto_total'],
        mode='markers',
        name='Montos de Certificados',
        marker=dict(size=15),
        customdata=df['id']  # Agregar el ID del certificado a customdata
    ))

    fig.update_layout(
        title='Certificados por Fecha de Envio y Monto',
        xaxis_title='Fecha de Envio',
        yaxis_title='Monto Total',
        xaxis=dict(
            tickformat='%d %b %Y',  # Formato para mostrar solo la fecha en formato amigable
            tickmode='array',
            tickvals=df['fecha_envio'].dt.date.unique(),  # Mostrar solo fechas únicas
            ticktext=[date.strftime('%d %b %Y') for date in df['fecha_envio'].dt.date.unique()]
        )
    )
    # Convertir el gráfico a JSON
    graph_json = pio.to_json(fig)
    return graph_json


def generar_grafico_cronograma(hechos_cronograma):
    # Leer los datos en un DataFrame de pandas
    df = pd.DataFrame(list(hechos_cronograma.values('detalle_id', 'fecha_programada', 'fecha_culminacion')))
    df.rename(columns={'detalle_id': 'detalle_id_cronograma'}, inplace=True)

    # Obtener los datos del modelo DetalleCronograma
    detalles_cronograma = DetalleCronograma.objects.all().values('id', 'detalle_id')
    df_detalles_cronograma = pd.DataFrame(list(detalles_cronograma))
    df_detalles_cronograma.rename(columns={'id': 'detalle_id_cronograma'}, inplace=True)

    # Obtener los datos del modelo Detalle
    detalles = Detalle.objects.all().values('id', 'rubro')
    df_detalles = pd.DataFrame(list(detalles))
    df_detalles.rename(columns={'id': 'detalle_id'}, inplace=True)
    # Realizar la unión entre los DataFrames de detalles_cronograma y detalles
    df_completo = df_detalles_cronograma.merge(df_detalles, on='detalle_id', how='left')
    # Realizar la unión entre el DataFrame original y el DataFrame completo
    df = df.merge(df_completo, on='detalle_id_cronograma', how='left')
    df.drop(['detalle_id'], axis=1, inplace=True)
    df.rename(columns={'detalle_id_cronograma': 'detalle_id'}, inplace=True)

    print(df)

    # Crear el gráfico de líneas
    fig = go.Figure()
    if not df.empty:
        # Añadir los puntos de las fechas programadas
        fig.add_trace(go.Scatter(
            x=df['fecha_programada'],
            y=df['detalle_id'],
            mode='markers',
            name='Fechas Programadas',
            text=df['rubro'],
            marker=dict(
                symbol='circle',
                color='blue',
                size=10,
                opacity=0.6
            ),
            hovertemplate='<b>Rubro:</b> %{text}<extra></extra>'
        ))

    # Filtrar solo las filas donde fecha_culminacion no es None
    df_culminacion = df.dropna(subset=['fecha_culminacion'])

    if not df_culminacion.empty:
        # Añadir los puntos de las fechas de culminación
        fig.add_trace(go.Scatter(
            x=df_culminacion['fecha_culminacion'],
            y=df_culminacion['detalle_id'],
            mode='markers',
            name='Fechas de Culminación',
            text=df_culminacion['rubro'],
            marker=dict(
                symbol='triangle-up',
                color=['orange' if x > y else 'green' if x == y else 'red' for x, y in zip(df_culminacion['fecha_programada'], df_culminacion['fecha_culminacion'])],
                size=10,
                opacity=0.6
            ),
            hovertemplate='<b>Rubro:</b> %{text}<extra></extra>'
        ))
    if fig.data:
        fig.update_layout(
            title='Cronograma de Actividades',
            xaxis_title='Fecha',
            yaxis_title='Rubro',
            yaxis=dict(
                type='category',  # Establecer el tipo de eje como categoría
                categoryorder='category ascending'
            ),
            xaxis=dict(
                tickformat='%d %b %Y',  # Formato para mostrar solo la fecha
                tickmode='array',
                tickvals=pd.concat([df['fecha_programada'], df['fecha_culminacion']]).dropna().unique(),
                # Mostrar solo fechas únicas
                ticktext=[date.strftime('%d %b %Y') for date in pd.concat([df['fecha_programada'], df['fecha_culminacion']]).dropna().unique()]
            ),
            legend=dict(x=0.1, y=1.1, orientation='h')
        )

    # Convertir el gráfico a JSON
    graph_json = pio.to_json(fig)
    return graph_json


def ver_resumen_obra(request, obra_id):
    obra = Obra.objects.get(id=obra_id)
    proyecto = obra.proyecto
    presupuesto = Presupuesto.objects.get(proyecto=proyecto)
    archivo_presupuesto = ArchivoPresupuesto.objects.get(presupuesto=presupuesto)
    cronograma = Cronograma.objects.get(archivo_presupuesto=archivo_presupuesto)
    nombre_obra = proyecto.nombre
    cliente = proyecto.cliente.nombre
    ing_obra = obra.encargado.first_name + ' ' + obra.encargado.last_name
    ing_presupuesto = presupuesto.encargado.first_name + ' ' + presupuesto.encargado.last_name
    pedidos = Pedido.objects.filter(obra=obra).order_by('-estado', 'fecha_solicitud')
    devoluciones = Devolucion.objects.filter(obra=obra).order_by('-estado', 'fecha_solicitud')
    # Obtener los datos de HechoMaterialPorObra



    hechos_cronograma = HechoCronograma.objects.filter(cronograma_id=cronograma.id)
    if hechos_cronograma:
        graph_json_cronograma = generar_grafico_cronograma(hechos_cronograma)
        print(graph_json_cronograma)
    else:
        graph_json_cronograma = None

    hechos_materiales = HechoMaterialPorObra.objects.filter(obra_id=obra.id)
    materiales_ids = [hecho.material_id for hecho in hechos_materiales]
    materiales = DimMaterial.objects.filter(material_id__in=materiales_ids)
    unidades_medida = {material.nombre: UnidadMedida.objects.get(id=material.medida_id).descripcion for material in materiales}

    graph_json = generar_grafico_materiales(hechos_materiales) if hechos_materiales.exists() else json.dumps({})

    hechos_certificados = HechoCertificado.objects.filter(presupuesto_id=presupuesto.id)
    certificados_json = generar_grafico_certificados(hechos_certificados) if hechos_certificados.exists() else json.dumps({})

    context = {
        'pedidos': pedidos,
        'nombre_obra': nombre_obra,
        'nombre_ing_obra': ing_obra,
        'nombre_ing_pres': ing_presupuesto,
        'nombre_cliente': cliente,
        'proyecto': proyecto,
        'obra': obra,
        'presupuesto': presupuesto,
        'hechos_materiales': hechos_materiales,
        'graph_json': graph_json,
        'graph_json_cronograma': graph_json_cronograma,
        'certificados_json': certificados_json,
        'unidades_medida': unidades_medida,
        'devoluciones': devoluciones
    }
    return render(request, 'pantallas_gerente/ver_resumen_obra.html', context)


def ver_pedido_gerente(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    materiales_pedido = MaterialPedido.objects.filter(
        pedido=pedido)  # Obtiene los materiales relacionados con el pedido
    return render(request, 'pantallas_gerente/ver_pedido_gerente.html', {'pedido': pedido, 'materiales_pedido': materiales_pedido})


def ver_ing_gerente(request, ingeniero_id=None):
    group_ingeniero = Group.objects.get(name='INGENIERO')
    usuarios_ingenieros = group_ingeniero.user_set.all()

    ingeniero_seleccionado = None
    obras = Obra.objects.all().select_related('proyecto', 'encargado', 'proyecto__cliente')
    presupuestos = Presupuesto.objects.all().select_related('proyecto', 'encargado', 'proyecto__cliente')
    cronograma_data = []
    certificado_data = []
    precision_cronograma_data = {'antes': 0, 'en_fecha': 0, 'despues': 0}
    cumplimiento_fecha_fin_data = {'antes': 0, 'en_fecha': 0, 'despues': 0}

    if ingeniero_id:
        ingeniero_seleccionado = get_object_or_404(User, id=ingeniero_id)
        obras = obras.filter(encargado=ingeniero_seleccionado)
        presupuestos = presupuestos.filter(encargado=ingeniero_seleccionado)


        for obra in obras:
            archivo_presupuesto = ArchivoPresupuesto.objects.filter(presupuesto__proyecto_id=obra.proyecto_id).first()
            if archivo_presupuesto:
                cronograma = Cronograma.objects.get(archivo_presupuesto=archivo_presupuesto)
                if cronograma:
                    total_actividades = HechoCronograma.objects.filter(cronograma_id=cronograma.id).count()
                    actividades_realizadas = HechoCronograma.objects.filter(cronograma_id=cronograma.id,
                                                                            fecha_culminacion__isnull=False).count()
                    porcentaje_actividades = (
                                                         actividades_realizadas / total_actividades) * 100 if total_actividades > 0 else 0
                    cronograma_data.append({
                        'obra': obra.proyecto.nombre,
                        'porcentaje_actividades': porcentaje_actividades,
                        'obra_id': obra.id
                    })
                    # Calcular precisión de cronograma
                    detalles_cronograma = HechoCronograma.objects.filter(cronograma_id=cronograma.id)
                    for detalle in detalles_cronograma:
                        if detalle.fecha_culminacion:
                            if detalle.fecha_culminacion < detalle.fecha_programada:
                                precision_cronograma_data['antes'] += 1
                            elif detalle.fecha_culminacion == detalle.fecha_programada:
                                precision_cronograma_data['en_fecha'] += 1
                            else:
                                precision_cronograma_data['despues'] += 1

        # Calcular porcentajes de montos del certificado
        for obra in obras:
            presupuesto = HechoPresupuesto.objects.filter(proyecto_id=obra.proyecto.id).first()
            if presupuesto:
                total_monto = float(presupuesto.monto_total)  # Convertir Decimal a float
                monto_aprobado = HechoCertificado.objects.filter(presupuesto_id=presupuesto.id, estado='A').aggregate(
                    total=Sum('monto_total'))['total']
                monto_aprobado = float(monto_aprobado) if monto_aprobado else 0
                porcentaje_monto = (monto_aprobado / total_monto) * 100 if total_monto > 0 else 0
                certificado_data.append({
                    'obra': obra.proyecto.nombre,
                    'porcentaje_monto': porcentaje_monto
                })
            # Calcular cumplimiento de fechas de fin
            if obra.estado == 'F' and obra.fecha_fin:
                fecha_estimada_fin = obra.fecha_inicio + timedelta(days=obra.plazo)
                if obra.fecha_fin < fecha_estimada_fin:
                    cumplimiento_fecha_fin_data['antes'] += 1
                elif obra.fecha_fin == fecha_estimada_fin:
                    cumplimiento_fecha_fin_data['en_fecha'] += 1
                else:
                    cumplimiento_fecha_fin_data['despues'] += 1

    obras_json = json.dumps(list(obras.values()), cls=CustomJSONEncoder)
    presupuestos_json = json.dumps(list(presupuestos.values()), cls=CustomJSONEncoder)
    context = {
        'ingenieros': usuarios_ingenieros,
        'obras': obras,
        'presupuestos': presupuestos,
        'obras_json': obras_json,
        'presupuestos_json': presupuestos_json,
        'ingeniero_seleccionado': ingeniero_seleccionado,
        'cronograma_data': json.dumps(cronograma_data),
        'certificado_data': json.dumps(certificado_data),
        'precision_cronograma_data': json.dumps(precision_cronograma_data),
        'cumplimiento_fecha_fin_data': json.dumps(cumplimiento_fecha_fin_data)
    }
    return render(request, 'pantallas_gerente/ver_ing_gerente.html', context)


class CustomJSONEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d')
        return super().default(obj)

def get_gestion_obras_presupuestos(request):
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    ingeniero_id = request.GET.get('ingeniero_id')
    page = request.GET.get('page', 1)

    items = []

    if tipo == 'Obra':
        if estado:
            items = Obra.objects.filter(estado=estado)
        else:
            items = Obra.objects.all()
        if ingeniero_id:
            items = items.filter(encargado_id=ingeniero_id)
        items = items.select_related('proyecto', 'encargado', 'proyecto__cliente')
    elif tipo == 'Presupuesto':
        if estado:
            items = Presupuesto.objects.filter(estado=estado)
        else:
            items = Presupuesto.objects.all()
        if ingeniero_id:
            items = items.filter(encargado_id=ingeniero_id)
        items = items.select_related('proyecto', 'encargado', 'proyecto__cliente')
    else:
        obras = Obra.objects.all()
        presupuestos = Presupuesto.objects.all()
        if ingeniero_id:
            obras = obras.filter(encargado_id=ingeniero_id)
            presupuestos = presupuestos.filter(encargado_id=ingeniero_id)
        items = list(obras) + list(presupuestos)

    paginator = Paginator(items, 5)
    page_obj = paginator.get_page(page)

    data = {
        'items': [
            {'id': item.id, 'nombre': item.proyecto.nombre, 'tipo': 'Obra' if isinstance(item, Obra) else 'Presupuesto',
             'estado': item.get_estado_display()} for item in page_obj],
        'estados': [{'valor': estado[0], 'nombre': estado[1]} for estado in
                    (Obra.ESTADOS if tipo == 'Obra' else Presupuesto.ESTADOS)] if tipo in ['Obra',
                                                                                           'Presupuesto'] else [],
    }
    return JsonResponse(data)



def get_estados_por_tipo(request):
    tipo = request.GET.get('tipo')
    estados = Obra.ESTADOS if tipo == 'Obra' else Presupuesto.ESTADOS
    return JsonResponse({'estados': [{'valor': estado[0], 'nombre': estado[1]} for estado in estados]})


def get_gestion_pedidos_devoluciones(request):
    tipo = request.GET.get('tipo')
    estado = request.GET.get('estado')
    ingeniero_id = request.GET.get('ingeniero_id')
    page = request.GET.get('page', 1)

    items = []

    if tipo == 'Pedido':
        if estado:
            items = Pedido.objects.filter(estado=estado)
        else:
            items = Pedido.objects.all()
        if ingeniero_id:
            items = items.filter(solicitante_id=ingeniero_id)
        items = items.select_related('obra', 'solicitante')
    elif tipo == 'Devolucion':
        if estado:
            items = Devolucion.objects.filter(estado=estado)
        else:
            items = Devolucion.objects.all()
        if ingeniero_id:
            items = items.filter(ingeniero_id=ingeniero_id)
        items = items.select_related('obra', 'ingeniero', 'pedido')
    else:
        pedidos = Pedido.objects.all()
        devoluciones = Devolucion.objects.all()
        if ingeniero_id:
            pedidos = pedidos.filter(solicitante_id=ingeniero_id)
            devoluciones = devoluciones.filter(ingeniero_id=ingeniero_id)
        items = list(pedidos) + list(devoluciones)

    paginator = Paginator(items, 5)
    page_obj = paginator.get_page(page)

    data = {
        'items': [
            {
                'id': item.id,
                'tipo': 'Pedido' if isinstance(item, Pedido) else 'Devolución',
                'obra': item.obra.proyecto.nombre,
                'estado': item.get_estado_display()
            } for item in page_obj],
        'estados': [{'valor': estado[0], 'nombre': estado[1]} for estado in
                    (Pedido.ESTADOS if tipo == 'Pedido' else Devolucion.ESTADOS)] if tipo in ['Pedido', 'Devolucion'] else [],
    }
    return JsonResponse(data)



def get_estados_por_tipo_pedido_devolucion(request):
    tipo = request.GET.get('tipo')
    estados = Pedido.ESTADOS if tipo == 'Pedido' else Devolucion.ESTADOS
    return JsonResponse({'estados': [{'valor': estado[0], 'nombre': estado[1]} for estado in estados]})


def get_obras(request):
    cliente_id = request.GET.get('cliente')
    ingeniero_id = request.GET.get('ingeniero')
    page_number = request.GET.get('page', 1)

    obras= Obra.objects.all().select_related('proyecto', 'encargado', 'proyecto__cliente')
    if cliente_id:
        obras= obras.filter(proyecto__cliente_id=cliente_id)
    if ingeniero_id:
        obras = obras.filter(encargado_id=ingeniero_id)

    paginator = Paginator(obras, 10)
    page_obj = paginator.get_page(page_number)

    obras_data = []
    for obra in page_obj:
        obras_data.append({
            'id': obra.id,
            'proyecto': {
                'nombre': obra.proyecto.nombre,
                'cliente': {
                    'nombre': obra.proyecto.cliente.nombre
                }
            },
            'encargado': {
                'get_full_name': obra.encargado.get_full_name()
            }
        })

    data = {
        'obras': obras_data,
        'page_obj': {
            'number': page_obj.number,
            'paginator': {
                'num_pages': paginator.num_pages,
            },
            'has_previous': page_obj.has_previous(),
            'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'has_next': page_obj.has_next(),
            'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        }
    }

    return JsonResponse(data)


def ver_obras_gerente(request):
    obras = Obra.objects.all().select_related('proyecto', 'encargado', 'proyecto__cliente')
    clientes_ids = obras.values_list('proyecto__cliente_id', flat=True).distinct()
    clientes = Cliente.objects.filter(id__in=clientes_ids)
    ingenieros_ids = obras.values_list('encargado_id', flat=True).distinct()
    ingenieros_encargados = User.objects.filter(id__in=ingenieros_ids)
    # Paginación
    paginator = Paginator(obras, 5)  # Mostrar 5 obras por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    cronograma_data, certificado_data = fetch_cronograma_certificado_data()

    context={
        'clientes': clientes,
        'ingenieros_encargados': ingenieros_encargados,
        'page_obj': page_obj,
        'certificado_data': json.dumps(certificado_data),
        'cronograma_data': json.dumps(cronograma_data)
    }
    return render(request, 'pantallas_gerente/ver_obras_gerente.html', context)


def fetch_cronograma_certificado_data(estado=None):
    cronograma_data = []
    certificado_data = []

    if not estado:
        obras = Obra.objects.all().select_related('proyecto', 'encargado', 'proyecto__cliente')
    elif estado == 'E':
        obras = Obra.objects.filter(estado='E').select_related('proyecto', 'encargado', 'proyecto__cliente')
    elif estado == 'F':
        obras = Obra.objects.filter(estado='F').select_related('proyecto', 'encargado', 'proyecto__cliente')
    else:
        obras = Obra.objects.none()
    print(obras)
    for obra in obras:
        archivo_presupuesto = ArchivoPresupuesto.objects.filter(presupuesto__proyecto_id=obra.proyecto_id).first()
        if archivo_presupuesto:
            try:
                cronograma = Cronograma.objects.get(archivo_presupuesto=archivo_presupuesto)
                total_actividades = HechoCronograma.objects.filter(cronograma_id=cronograma.id).count()
                actividades_realizadas = HechoCronograma.objects.filter(cronograma_id=cronograma.id, fecha_culminacion__isnull=False).count()
                porcentaje_actividades = (actividades_realizadas / total_actividades * 100) if total_actividades > 0 else 0
                cronograma_data.append({
                    'obra': obra.proyecto.nombre,
                    'porcentaje_actividades': porcentaje_actividades,
                    'obra_id': obra.id
                })
            except Cronograma.DoesNotExist:
                pass

        presupuesto = HechoPresupuesto.objects.filter(proyecto_id=obra.proyecto.id).first()
        if presupuesto:
            total_monto = float(presupuesto.monto_total)  # Convertir Decimal a float
            monto_aprobado = HechoCertificado.objects.filter(presupuesto_id=presupuesto.id, estado='A').aggregate(total=Sum('monto_total'))['total']
            monto_aprobado = float(monto_aprobado) if monto_aprobado else 0
            porcentaje_monto = (monto_aprobado / total_monto * 100) if total_monto > 0 else 0
            certificado_data.append({
                'obra': obra.proyecto.nombre,
                'porcentaje_monto': porcentaje_monto
            })

    return cronograma_data, certificado_data


def get_cronograma_certificado_data(request, estado=None):
    cronograma_data, certificado_data = fetch_cronograma_certificado_data(estado)
    data = {
        'certificado_data': certificado_data,
        'cronograma_data': cronograma_data
    }
    return JsonResponse(data, safe=False)
